import datetime
import openpyxl

from application.agent.models import Store, PromoCode
from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.utils.translation import ugettext_lazy as _
from django.views.generic import View
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


class ReportStoresView(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportStoresView, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        def save_to_cell(ws, row, column, value):

            font = Font(name='Arial', size=8)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            cell = ws.cell(row, column)
            cell.value = value
            cell.font = font
            cell.border = border

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="loyalty_stores_%s.xls"' % datetime.datetime.now()

        # Init values
        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Loyalty stores'))

        # Title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        cell = ws['A2']
        cell.value = str(_('Loyalty stores {date_start} {date_end}').format(
            date_start='с ' + datetime_start.strftime('%d.%m.%Y') if datetime_start else '',
            date_end='по ' + datetime_end.strftime('%d.%m.%Y') if datetime_end else '',
        ))
        cell.font = Font(name='Arial', size=14, bold=True)
        rd = ws.row_dimensions[2]
        rd.height = 18

        # Stores
        promo_code = PromoCode.objects.filter(store=OuterRef('pk'), is_used=True).values('code')[:1]
        items = Store.objects.filter(loyalty_program__isnull=False).annotate(promo_code=Subquery(promo_code)).\
            prefetch_related('city', 'user', 'loyalty_program', 'loyalty_department')
        if datetime_start:
            items = items.filter(date_create__gte=datetime_start)
        if datetime_end:
            items = items.filter(date_create__lte=datetime_end)
        if request.GET.get('department'):
            items = items.filter(loyalty_department=request.GET.get('department'))
        if request.GET.get('program'):
            items = items.filter(loyalty_program=request.GET.get('program'))

        # Настройки
        start_row = 4
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        header = (
            ('Ид', 6),
            ('Дата регистрации', 14),
            ('Название', 26),
            ('Контакт', 20),
            ('Телефон', 10),
            ('Адрес', 34),
            ('ИНН', 10),
            ('Регион', 12),
            ('Отдел', 12),
            ('Программа', 16),
            ('1c код', 10),
            ('План', 8),
            ('Факт', 8),
            ('Кэш-бэк заработан', 15),
            ('Кэш-бэк выплачен', 15),
            ('Кэш-бэк к выплате', 15),
            ('Текущая задолженность', 21),
            ('Просроченная задолженность', 26),
            ('ID пользователя', 15),
            ('ФИО пользователя', 17),
            ('Дата регистрации пользователя', 27),
            ('E-mail пользователя', 26),
            ('Рекомендатель', 15),
            ('Торговый представитель', 22),
            ('Промо код', 20),
        )

        # Rows
        row = start_row
        for i in items:

            row += 1

            rd = ws.row_dimensions[row]
            rd.height = 13

            save_to_cell(ws, row, 1, i.id)
            save_to_cell(ws, row, 2, i.date_create)
            save_to_cell(ws, row, 3, i.name)
            save_to_cell(ws, row, 4, i.contact)
            save_to_cell(ws, row, 5, i.phone)
            save_to_cell(ws, row, 6, i.address)
            save_to_cell(ws, row, 7, i.inn)
            save_to_cell(ws, row, 8, i.city.name if i.city else '')
            save_to_cell(ws, row, 9, i.loyalty_department.name if i.loyalty_department else '')
            save_to_cell(ws, row, 10, i.loyalty_program.name if i.loyalty_program else '')
            save_to_cell(ws, row, 11, i.loyalty_1c_code)
            save_to_cell(ws, row, 12, i.loyalty_plan)
            save_to_cell(ws, row, 13, i.loyalty_fact)
            save_to_cell(ws, row, 14, i.loyalty_sumcashback)
            save_to_cell(ws, row, 15, i.loyalty_cashback_payed)
            save_to_cell(ws, row, 16, i.loyalty_cashback_to_pay)
            save_to_cell(ws, row, 17, i.loyalty_debt)
            save_to_cell(ws, row, 18, i.loyalty_overdue_debt)
            save_to_cell(ws, row, 19, i.user.id)
            save_to_cell(ws, row, 20, i.user.fio)
            save_to_cell(ws, row, 21, i.user.date_join)
            save_to_cell(ws, row, 22, i.user.email)
            save_to_cell(ws, row, 23, i.user.advisor)
            save_to_cell(ws, row, 24, i.agent)
            if request.user.email in settings.PROMO_CODES_USERS:
                save_to_cell(ws, row, 25, i.promo_code)
            else:
                save_to_cell(ws, row, 25, 'Нет прав')

        # Filters
        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='f4ecc5', end_color='f4ecc5', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), items.count() + start_row)
        ws.auto_filter.ref = filter_range

        response.content = save_virtual_workbook(wb)
        return response
