import json
import requests
import openpyxl
import os

from datetime import datetime
from django.conf import settings
from django.core.files import File
from django.core.mail import EmailMessage
from io import BytesIO
from django.db.models import OuterRef, Subquery, Exists

from application.survey.models import User, UserDeviceIceman
from .models import StoreTask, Order


def push(key, title, message, category):

    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'key=AAAAX1VoAFI:APA91bFySaOvzLqLfgCVJPu94_vl6pWzVluRFififGcXEWnufDU7VBupdOhMrudAOt7XvnVYn5S76'
                         'E07bSMsnIJkTpVtkXYTcUR9ytD6OyuKimVUl5z8PdJ0RcjF7bTXC9nJHYeSnHWX'
    }

    to = key

    if key is None:
        to = '/topics/all'

    data = {
        'to': to,
        'notification': {
            'title': title,
            'body': message,
            'sound': 'default',
            'priority': 'high',
        },
        'data': {
            'click_action': 'FLUTTER_NOTIFICATION_CLICK',
            'category': category,
        },
    }

    try:
        result = requests.post('https://fcm.googleapis.com/fcm/send', headers=headers, data=json.dumps(data))
    except Exception as e:
        return f'Ошибка: {e}'

    return result.text


def sync_stocks(data, source):

    from application.iceman.models import Stock, Source

    try:
        source_obj = Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # Добавляем/обновляем склады
    for item in data['sklad']:

        sys_name = item['id']
        name = item['name']

        try:
            stock_obj = Stock.objects.get(sys_name=sys_name)
        except Stock.DoesNotExist:
            stock_obj = Stock(sys_name=sys_name)

        stock_obj.source = source_obj
        stock_obj.name = name
        stock_obj.save()

    return 'Ok'


def sync_stores(data, source, data_name='Clients', is_order_task=True, store_type='iceman'):

    from application.iceman.models import Region, Source, Store, StoreStock, Stock
    from django.db.models import Q

    try:
        source_obj = Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # Добавляем/обновляем магазины
    for item in data[data_name]:

        code = item['id']
        name = item['name']
        is_agreement = item['dogovor'] == 'Да'
        lpr_fio = item['contact']
        lpr_phone = item['telephone']
        inn = item['INN']
        schedule = item['dostavka']
        provod = bool(item['provod']) if item.get('provod') is not None else None

        try:
            id_iceman = int(str(item['id_iceman']).replace('ICMS', ''))
        except:
            id_iceman = None

        try:
            payment_days = int(item['delay'])
        except:
            payment_days = None
        try:
            sum_restrict = int(item['credit'])
        except:
            sum_restrict = 0

        region_name = item['region']
        region = None
        if region_name.strip() != '':
            try:
                region = Region.objects.filter(Q(name_1c__iexact=region_name) | Q(name_1c_2__iexact=region_name)).first()
            except Region.DoesNotExist:
                pass

        address = item['adress']

        if address.strip() == '':
            continue

        price_type = item['PriceType']

        # Магазин по ИД
        store_obj = None
        if id_iceman:
            try:
                store_obj = Store.objects.get(id=id_iceman)

                # Будем обновлять координаты если пришел новый адрес в существующем магазине
                if store_obj.address != address:
                    store_obj.auto_coord = True

                # Обновлям код
                store_obj.code = code

            except Store.DoesNotExist:
                pass

        # Магаз по коду
        if store_obj is None:
            try:
                store_obj = Store.objects.get(code=code)

                # Будем обновлять координаты если пришел новый адрес в существующем магазине
                if store_obj.address != address:
                    store_obj.auto_coord = True

            except Store.DoesNotExist:
                store_obj = Store(code=code)

        schedule_items = schedule.split(';')
        new_schedule_items = []
        for schedule_item in schedule_items:
            if schedule_item.lower() == 'понедельник':
                new_schedule_items.append('1')
            if schedule_item.lower() == 'вторник':
                new_schedule_items.append('2')
            if schedule_item.lower() == 'среда':
                new_schedule_items.append('3')
            if schedule_item.lower() == 'четверг':
                new_schedule_items.append('4')
            if schedule_item.lower() == 'пятница':
                new_schedule_items.append('5')
            if schedule_item.lower() == 'суббота':
                new_schedule_items.append('6')
            if schedule_item.lower() == 'воскресенье':
                new_schedule_items.append('7')
        schedule = ','.join(new_schedule_items)

        store_obj.source = source_obj
        store_obj.name = name
        store_obj.is_agreement = is_agreement
        store_obj.lpr_fio = lpr_fio
        store_obj.lpr_phone = lpr_phone
        store_obj.inn = inn
        store_obj.address = address
        if provod is not None:
            store_obj.is_entry = provod
        if is_order_task:
            store_obj.price_type = 'Цена2'
        else:
            store_obj.price_type = price_type
        store_obj.sum_restrict = sum_restrict
        if region:
            store_obj.region = region
        if payment_days:
            store_obj.payment_days = payment_days
        store_obj.schedule = schedule
        store_obj.is_order_task = is_order_task
        store_obj.type = store_type
        store_obj.save()

        # Координаты
        if store_obj.address and (not store_obj.longitude or not store_obj.latitude):
            store_obj.auto_coord = True
            store_obj.save()

        # ИНН
        if store_obj.inn and not store_obj.inn_name:
            store_obj.inn_auto = True
            store_obj.save()
            if store_obj.inn_region and store_obj.region is None:
                region = Region.objects.filter(
                    Q(short_name__iexact=store_obj.inn_region) | Q(short_name_2__iexact=store_obj.inn_region)
                ).first()
                if region is None:
                    region = Region(name=store_obj.inn_region, short_name=store_obj.inn_region)
                    region.save()
                store_obj.region = region
                store_obj.save()

        # Склады
        stock = item['sklad']

        # Пустой склад, ищем по региону или берем дефолт
        if not stock or stock == '00000000-0000-0000-0000-000000000000':
            if store_obj.region and store_obj.region.stock:
                stock_obj = store_obj.region.stock
            else:
                stock_obj = Stock.objects.filter(default=True, source=source_obj).first()
        # Склад есть, ищем в таблице или слоздаем новый
        else:
            try:
                stock_obj = Stock.objects.get(sys_name=stock)
            except Stock.DoesNotExist:
                stock_obj = Stock(sys_name=stock, source=source_obj, name='Неизвестный склад из 1с')
                stock_obj.save()

        if stock_obj is not None:

            # Добавляем склад
            store_stock = StoreStock.objects.filter(store=store_obj, stock=stock_obj).exists()
            if not store_stock:
                StoreStock(store=store_obj, stock=stock_obj).save()

            # Удаляем старые склады
            store_stocks = StoreStock.objects.filter(store=store_obj).exclude(stock=stock_obj)
            if store_stocks:
                store_stocks.delete()

    return 'Ok'


def sync_goods(data, source):

    from application.iceman.models import Source, Product, SourceProduct, Brand, Category

    try:
        source_obj = Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # SourceProduct.objects.filter(source=source_obj).update(is_updated=False)

    # Добавляем/обновляем товары
    for item in data['tovar']:

        # Получаем поля
        code = item['code']
        unit = item['unit']
        price = item['price']
        embeddability = item['embeddability']
        min = item['min']
        barcode = item['barcode']
        weight = item['weight']
        brand_name = item['manufacturer']
        name = item['name']
        category_name = item['groupdop']

        # Продукт
        try:
            product = Product.objects.get(barcode=barcode)
        except Product.DoesNotExist:
            product = Product(barcode=barcode)

        # Производитель
        try:
            brand = Brand.objects.get(name=brand_name)
        except Brand.DoesNotExist:
            brand = Brand(name=brand_name)
            brand.save()

        # Раздел
        try:
            category = Category.objects.get(name=category_name)
        except Category.DoesNotExist:
            category = Category(name=category_name)
            category.save()

        product.name = name
        product.brand = brand
        product.code = code
        product.unit = unit
        product.box_count = embeddability
        product.min_count = min
        product.price = round(price, 2)
        product.barcode = barcode
        product.weight = weight * 1000
        product.save()

        # Раздел
        if not category.product_set.filter(pk=product.pk).exists():
            product.categories.add(category)
            product.save()

        # Изображение
        if item['picture'] != '' and not product.image:
            io = BytesIO()
            response = requests.get(item['picture'])
            if response.status_code == 200:
                io.write(response.content)
                extension = item['picture'].split('.')[-1]
                product.image.save(f'image.{extension}', File(io))

        # Продукт в магазине
        try:
            sp_obj = SourceProduct.objects.get(product=product, source=source_obj)
        except SourceProduct.DoesNotExist:
            sp_obj = SourceProduct(product=product, source=source_obj)
        sp_obj.is_updated = True
        sp_obj.unit = unit
        sp_obj.box_count = embeddability
        sp_obj.min_count = min
        sp_obj.price = round(price, 2)
        sp_obj.save()

    # Удаляем продукты в магазине
    # SourceProduct.objects.filter(is_updated=False).delete()

    return 'Ok'


def sync_prices(data, source):

    from application.iceman.models import Source, Product, SourceProduct, SourceProductPrice

    try:
        source_obj = Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # Добавляем/обновляем цены
    for item in data['cen']:

        # Получаем товар
        products = Product.objects.filter(code=item['code'])

        for product in products:

            # Получаем товар в магазине
            try:
                source_product = SourceProduct.objects.get(product=product, source=source_obj)
            except SourceProduct.DoesNotExist:
                continue

            # Получаем запись
            try:
                product_price = SourceProductPrice.objects.get(product=source_product, price_type=item['type_cost'])
            except SourceProductPrice.DoesNotExist:
                product_price = SourceProductPrice(product=source_product, price_type=item['type_cost'])

            product_price.price = item['cost']
            product_price.save()

    return 'Ok'


def sync_stocks_goods(data, source):

    from application.iceman.models import Source, Product, ProductStock, Stock

    try:
        source_obj = Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # Добавляем/обновляем цены
    for item in data['Ostatok']:

        # Получаем товар
        products = Product.objects.filter(code=item['id_tovar'])

        for product in products:

            # Получаем склад
            try:
                stock = Stock.objects.get(sys_name=item['id_sklad'])
            except Stock.DoesNotExist:
                continue

            # Получаем запись
            try:
                obj = ProductStock.objects.get(product=product, stock=stock)
            except ProductStock.DoesNotExist:
                obj = ProductStock(product=product, stock=stock)

            obj.count = item['ostatok']
            obj.save()

    return 'Ok'


def excel_get_string(val, max_length=None):

    if type(val) == float:
        val = str(val).replace('.0', '')

    if type(val) == int:
        val = str(val).replace('.', '')

    val = str(val)

    if max_length is not None:
        val = val[:max_length]

    if val == 'nan':
        return None

    return str(val)


def make_order_excel(order, mail_send=True):

    # Открываем шаблон
    excel_file = os.path.join(settings.DATA_DIR, 'iceman_order.xlsx')

    # Новый файл
    openpyxl.load_workbook(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.worksheets[0]

    # Заполняем данными
    ws['C1'] = f'ЗАКАЗ № {order.order_id}'
    ws['D2'] = order.store.inn_name if order.store is not None else ''
    ws['D3'] = order.store.lpr_phone if order.store is not None else ''
    ws['D4'] = order.store.inn if order.store is not None else ''
    ws['D5'] = order.store.inn_kpp if order.store is not None else ''
    ws['D7'] = order.store.lpr_fio if order.store is not None else ''
    ws['D8'] = datetime.today().strftime('%d.%m.%Y')
    ws['D9'] = order.delivery_date.strftime('%d.%m.%Y')
    ws['D10'] = order.delivery_address
    ws['D11'] = order.comment

    source = order.store.source if order.store is not None else None

    ws['G2'] = source.partner_name if source is not None else ''
    ws['G3'] = source.partner_email if source is not None else ''
    ws['G4'] = source.partner_fio if source is not None else ''
    ws['G5'] = source.partner_phone if source is not None else ''

    ws['G7'] = order.user.id if order.user is not None else ''
    ws['G8'] = order.user.fio if order.user is not None else ''
    ws['G9'] = order.user.phone if order.user is not None else ''
    ws['G10'] = order.store.id if order.store is not None else ''

    ws['I2'] = f'Отсрочка дней: {order.payment_days}'
    ws['I3'] = order.get_payment_type_display()
    ws['I4'] = order.get_payment_method_display()
    ws['I5'] = ('Да' if order.store.is_agreement else 'Нет') if order.store is not None else ''
    ws['I6'] = order.user.advisor if order.user is not None else ''

    # Товары
    row = 13
    count = 0
    for product in order.products:
        ws[f'B{row}'] = f'{row - 12}'
        ws[f'C{row}'] = product.name
        ws[f'D{row}'] = product.unit
        ws[f'E{row}'] = product.weight
        ws[f'F{row}'] = product.box_count
        ws[f'G{row}'] = product.count
        ws[f'H{row}'] = round(product.price_one, 2)
        ws[f'I{row}'] = round(product.price, 2)
        count += product.count
        row += 1

    ws.delete_rows(row, 70 - order.products.count())

    ws[f'G{row}'] = count
    ws[f'I{row}'] = order.price

    # Ссылки на документы
    row += 3
    if order.store is not None:
        for document in order.store.documents:
            ws[f'C{row}'] = f'{document.get_type_display()}, страница {document.number}'
            link = f'https://admin.shop-survey.ru{document.file.url}'
            ws[f'D{row}'] = f'=HYPERLINK("{link}", "{link}")'
            row += 1

    # Сохраняем
    file_name = f'/tmp/order_{order.id}.xlsx'
    wb.save(file_name)

    # Отправляем на почту
    if mail_send and order.store is not None and order.store.source is not None and order.store.source.email is not None and \
            order.store.source.email != '':
        try:
            msg = EmailMessage(
                f'Заказ № {order.order_id}',
                f'Заказ № {order.order_id}',
                settings.EMAIL_FROM,
                order.store.source.email.split(';')
            )
            msg.content_subtype = 'html'
            msg.attach_file(file_name)
            msg.send()
        except:
            pass


def sync_orders(data, source):

    from application.iceman.models import Source, Order
    from application.survey.models import TasksExecution

    try:
        Source.objects.get(sys_name=source)
    except Source.DoesNotExist:
        raise Exception(f'Источник с системным именем {source} не найден')

    # Добавляем/обновляем товары
    for item in data['zakaz']:

        order_id = str(item['doc']).replace('ICM', '')
        sum = item['sum']
        pay = item['pay']
        status = item['status']
        payment_url = item['payment_link']

        try:
            delivery_date = datetime.strptime(item['date'], '%Y-%m-%dT%H:%M:%S')
        except:
            delivery_date = None

        try:
            order_id = int(order_id)
        except ValueError:
            continue

        try:
            order_obj = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            continue

        # Обновилась сумма
        if order_obj.payment_sum_correct != sum:

            order_obj.payment_sum_correct = sum
            order_obj.payment_sum = pay
            order_obj.payment_sum_user = order_obj.user_sum

            if order_obj.task_id is not None:
                try:
                    te_obj = TasksExecution.objects.get(id=order_obj.task_id)
                    te_obj.money = order_obj.user_sum
                    te_obj.save(update_fields=['money'])
                except TasksExecution.DoesNotExist:
                    pass

        # Обновляем
        order_obj.payment_sum_correct = sum
        order_obj.payment_sum = pay
        order_obj.payment_url = payment_url

        # Дата
        if delivery_date:
            order_obj.delivery_date = delivery_date

        # Заказ отменен
        if status == 'x':
            order_obj.status = 4
        if status == 'a' and order_obj.status == 4:
            order_obj.status = 1

        # Синхронизация с 1с
        order_obj.sync_1c = True
        order_obj.sync_1c_date = datetime.now()

        order_obj.save()

    return 'Ok'


def message_reciver_filter(request, get_func):

    full_filter = {}
    region = None

    id_user = get_func('id_user')
    if id_user:
        full_filter['id'] = id_user
    else:

        if get_func('advisor'):
            full_filter['advisor__contains'] = get_func('advisor')

        status_legal = get_func('status_legal')
        if status_legal and (status_legal != 'all'):
            full_filter['status_legal'] = status_legal

        user_type = get_func('user_type')
        if user_type and (user_type != 'all'):
            full_filter['type'] = user_type

        if get_func('iceman_status'):
            full_filter['status_iceman'] = get_func('iceman_status')

    users_list = User.objects.filter(**full_filter)

    if not id_user:
        region = get_func('region')
        if region:  
            region_users = StoreTask.objects.filter(
                only_user_id=OuterRef('id'),
                region=region
            )

            users_list = users_list.annotate(
                region_users=Exists(region_users)).filter(region_users=True
            )

        last_order_data_filter = {}
        last_order_data = get_func('last_order_date')
        if last_order_data:
            if isinstance(last_order_data, str):
                last_order_data_dt = datetime.strptime(last_order_data, '%d.%m.%Y')
            else:
                last_order_data_dt = last_order_data
            users_without_orders = Order.objects.filter(
                user__id=OuterRef('id'),
                date_create__lt=last_order_data_dt
            ).order_by('user__id').distinct('user__id').values('user__id')
            last_order_data_filter['id__in'] = Subquery(users_without_orders)

            users_list = users_list.filter(**last_order_data_filter)
        
        users_overdue_filter = {}
        is_overdue = get_func('is_overdue')
        if is_overdue and is_overdue == 'yes':
            users_overdues = [
                order.user.id for order in Order.objects.select_related('user').all()
                if order.days_overdue and (order.days_overdue > 0)
            ]
            users_overdue_filter['id__in'] = users_overdues
        elif is_overdue and is_overdue == 'no':
            users_overdues = [
                order.user.id for order in Order.objects.select_related('user').all()
                if order.user and (order.days_overdue is None)
            ]
            users_overdue_filter['id__in'] = users_overdues

        users_list = users_list.filter(**users_overdue_filter)

    users_list = users_list.filter(id__in=UserDeviceIceman.objects.values('user__id'))

    if (not full_filter and
        not region and
        not last_order_data_filter and
        not users_overdue_filter):
        is_empty_filter = True
    else:
        is_empty_filter = False

    return users_list, is_empty_filter
