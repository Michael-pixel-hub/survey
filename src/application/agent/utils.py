import datetime
import hashlib
import re
import os
import openpyxl
import sys
import xlrd

from application.iceman.models import Order as IcemanOrder
from django.conf import settings
from django.core.cache import cache
from django.core.mail import EmailMessage
from django.db import transaction
from django.utils.translation import ugettext_lazy as _

from preferences.utils import get_setting

from .dadata import Dadata


def import_data_from_file(file, file_name):

    from .models import Import, Good, Category, Brand, City, GoodPrice

    import_obj = Import(file_name=file_name)
    import_obj.save()

    try:

        rb = xlrd.open_workbook(file, formatting_info=True)
        sheet = rb.sheet_by_index(0)
        import_obj.rows_count = sheet.nrows
        import_obj.save()

        row = sheet.row_values(4)
        row = [item.lower() for item in row]

        code_index = row.index('код')
        category_index = row.index('группа')
        good_index = row.index('номенклатура')
        brand_index = row.index('марка (бренд)')
        unit_index = row.index('ед.изм.')
        box_count_index = row.index('затарка')
        try:
            min_count_index = row.index('минимальное количество')
        except:
            min_count_index = -1

        regions = City.objects.all()
        price_indexes = []
        first_price_index = sys.maxsize
        for city in regions:
            try:
                index = row.index('цена %s' % city.code.lower())
                price_indexes.append((city, index))
                if index < first_price_index:
                    first_price_index = index
            except ValueError:
                pass

        try:
            price_index = row.index('цена')
        except:
            if first_price_index == sys.maxsize:
                raise Exception(_('Column with price not found'))

            price_index = first_price_index

        for row_num in range(5, sheet.nrows):

            if cache.get('chl_bot_import_cancel'):
                cache.delete('chl_bot_import_cancel')
                import_obj.status = 3
                import_obj.save()
                return

            xfx = sheet.cell_xf_index(row_num, 2)
            xf = rb.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            pattern_colour = rb.colour_map[bgx]

            row = sheet.row_values(row_num)

            code = row[code_index] if code_index > -1 else ''
            code = str(code).strip()[:100]
            code = re.sub(' +', ' ', code).strip()

            category = row[category_index] if category_index > -1 else ''
            category = str(category).strip()[:200]
            category = re.sub(' +', ' ', category).strip().capitalize()
            category = category.capitalize()

            brand = row[brand_index] if brand_index > -1 else ''
            brand = str(brand).strip()[:200]
            brand = re.sub(' +', ' ', brand).strip()
            brand = brand.capitalize()

            good = row[good_index] if good_index > -1 else ''
            good = str(good).strip()[:100]
            good = re.sub(' +', ' ', good)
            good = good.replace('*', '').strip()

            unit = row[unit_index] if unit_index > -1 else ''
            unit = str(unit).strip()[:100]
            unit = re.sub(' +', ' ', unit).strip()

            price = row[price_index] if price_index > -1 else ''
            price = str(price).strip()[:100]
            price = re.sub(' +', ' ', price).strip()

            box_count = row[box_count_index] if box_count_index > -1 else ''
            box_count = str(box_count).strip()

            min_count = row[min_count_index] if min_count_index > -1 else ''
            min_count = str(min_count).strip()

            try:
                price = round(float(price), 2)
            except:
                price = 0

            try:
                box_count = int(round(float(box_count)))
            except:
                box_count = 0

            try:
                min_count = int(round(float(min_count)))
            except:
                min_count = 0

            if not code or not category or not brand or not good or not unit:
                continue

            category_obj, created = Category.objects.get_or_create(name=category)
            brand_obj, created = Brand.objects.get_or_create(name=brand)

            try:
                good_obj = Good.objects.get(code=code)
            except Good.DoesNotExist:
                good_obj = Good(code=code)

            good_obj.name = good
            good_obj.unit = unit
            good_obj.price = price
            good_obj.category = category_obj
            good_obj.brand = brand_obj
            if box_count:
                good_obj.box_count = box_count
            if min_count:
                good_obj.min_count = min_count
            if pattern_colour == (255, 255, 0):
                good_obj.is_popular = True
            else:
                good_obj.is_popular = False
            if pattern_colour == (51, 204, 204):
                good_obj.is_oeskimo = True
            else:
                good_obj.is_oeskimo = False
            good_obj.save()

            if price_indexes:
                gp_exists = [{'gp': gp, 'updated': False} for gp in GoodPrice.objects.filter(good=good_obj)]
                for pi in price_indexes:
                    price_region = row[pi[1]]
                    price_region = str(price_region).strip()[:100]
                    price_region = re.sub(' +', ' ', price_region).strip()
                    try:
                        price_region = round(float(price_region), 2)
                    except:
                        price_region = 0
                    if price_region > 0:
                        try:
                            gp = GoodPrice.objects.get(good=good_obj, city=pi[0])
                            if gp.price != price_region:
                                gp.price = price_region
                                gp.save()
                            for gpe in gp_exists:
                                if gp.id == gpe['gp'].id:
                                    gpe['updated'] = True
                        except GoodPrice.DoesNotExist:
                            gp = GoodPrice(good=good_obj, city=pi[0])
                            gp.price = price_region
                            gp.save()
                for gpe in gp_exists:
                    if not gpe['updated']:
                        gpe['gp'].delete()

            try:
                Import.objects.get(status=1)
            except Import.DoesNotExist:
                return

            import_obj.rows_process = row_num + 1
            import_obj.save()

        import_obj.status = 2
        import_obj.date_end = datetime.datetime.now()
        import_obj.save()

    except Exception as e:

        import_obj.status = 4
        import_obj.report_text = '%s!\n\n%s' % (str(_('Error')), str(e))
        import_obj.save()

        return


def make_order_excel(order):

    from application.agent.models import OrderGood

    goods = list(OrderGood.objects.filter(order=order).prefetch_related('store'))
    stores = []

    for i in goods:
        is_store = False
        for j in stores:
            if j.id == i.store.id:
                is_store = True
        if not is_store:
            stores.append(i.store)

    files = []
    store_idx = 0

    for i in stores:

        store_idx += 1

        excel_file = os.path.join(settings.DATA_DIR, 'order.xlsx')

        openpyxl.load_workbook(excel_file)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.worksheets[0]

        ws['C2'] = order.id
        ws['L2'] = order.id
        ws['C3'] = i.name
        ws['C4'] = i.inn
        ws['C5'] = order.user.advisor
        ws['C6'] = order.delivery_date
        ws['C7'] = order.delivery_date
        ws['G5'] = order.delivery_address
        ws['L3'] = order.user.id
        ws['L4'] = i.id

        goods_count = 0
        goods_count_all = 0
        goods_sum = 0

        row = 9

        for g in goods:

            if i.id != g.store.id:
                continue

            goods_count += 1
            goods_count_all += g.count
            goods_sum += g.count * g.price

            row = 8 + goods_count

            ws.cell(row=row, column=1).value = goods_count
            ws.cell(row=row, column=3).value = g.code
            ws.cell(row=row, column=4).value = g.name
            ws.cell(row=row, column=5).value = g.unit
            ws.cell(row=row, column=6).value = g.count
            ws.cell(row=row, column=10).value = g.price
            ws.cell(row=row, column=12).value = g.price * g.count

        for s in range(0, 152-goods_count):
            ws.delete_rows(row + 1, 1)

        ws.cell(row=row + 1, column=3).value = goods_count
        ws.cell(row=row + 2, column=3).value = goods_count_all
        ws.cell(row=row + 3, column=3).value = goods_sum

        file_name = '/tmp/order_%s_%s.xlsx' % (order.id, store_idx)
        wb.save(file_name)
        files.append(file_name)

    msg = EmailMessage('Заказ №%s из телегерам бота @Serveyor_bot' % order.id, 'Заказ №%s' % order.id,
                       settings.EMAIL_FROM, get_setting('agent_orderemail').split(';'))
    msg.content_subtype = "html"
    for f in files:
        msg.attach_file(f)
        msg.send()

    for f in files:
        try:
            os.rename(f, os.path.join(settings.SALES_PATH, '%s.xlsx' % order.id))
        except:
            pass
        # try:
        #     os.unlink(f)
        # except:
        #     pass

    return 'Order %s mail sent' % order.id


def send_user_status_change(obj, old_status):

    from application.telegram.models import String
    from application.telegram.tasks import send_message

    comment = 'Нет'
    if obj.comments_status:
        comment = obj.comments_status

    store_s = obj.store.name if obj.store else ''

    s = String.get_string('message_agent_order_status_changed').format(
        order=obj.id,
        sum=obj.sum,
        date=obj.date_order.strftime('%d.%m.%Y'),
        store=store_s,
        status=obj.get_status_display(),
        comment=comment,
        old_status=old_status,
    )

    send_message.delay(obj.user.telegram_id, s)


def check_inn(store_id):

    from application.agent.models import Store

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return 'Store does not exists'

    api_key = get_setting('agent_dadataapikey')
    secret_key = get_setting('agent_dadatasecretkey')

    if not api_key:
        return 'No api key'

    if not secret_key:
        return 'No secret key'

    dadata = Dadata(api_key, secret_key)
    data = dadata.find_by_id('party', store.inn)

    return data

    # print(data)
    #
    # for k, v in data[0].items():
    #     print(k, v)


def save_inn(store_id):

    from application.agent.models import Store

    data = check_inn(store_id)

    if type(data) == str:
        return data

    if data is None:
        return 'No data'

    if len(data) < 1:
        return 'No data'

    try:
        store = Store.objects.get(id=store_id)
    except Store.DoesNotExist:
        return 'Store does not exists'

    try:
        store.inn_name = data[0]['value']
    except:
        pass

    try:
        store.inn_full_name = data[0]['data']['name']['full_with_opf']
    except:
        pass

    try:
        store.inn_director_title = data[0]['data']['management']['post']
    except:
        pass

    try:
        store.inn_director_name = data[0]['data']['management']['name']
    except:
        pass

    try:
        store.inn_address = data[0]['data']['address']['value']
    except:
        pass

    try:
        store.inn_region = data[0]['data']['address']['data']['region']
    except:
        pass

    try:
        store.inn_kpp = data[0]['data']['kpp']
    except:
        pass

    if not store.inn_director_name:
        try:
            if data[0]['data']['type'] == 'INDIVIDUAL':
                store.inn_director_name = data[0]['data']['name']['full']
                store.inn_director_title = data[0]['data']['opf']['full']
        except:
            pass

    try:
        store.inn_ogrn = data[0]['data']['ogrn']
    except:
        pass

    try:
        store.inn_okved = data[0]['data']['okved']
    except:
        pass

    store.save()
    return 'Ok'


def agent_1c_goods_prices(data):

    from application.agent.models import GoodPriceType, Good

    GoodPriceType.objects.all().update(is_processed=False)

    for item in data['goodsPrice']:

        try:
            good = Good.objects.get(code=item['code'])
        except Good.DoesNotExist:
            continue

        try:
            good_price_obj = GoodPriceType.objects.get(good=good, price_type=item['PriceType'])
        except GoodPriceType.DoesNotExist:
            good_price_obj = GoodPriceType(good=good, price_type=item['PriceType'])

        good_price_obj.price = round(item['price'], 2)
        good_price_obj.is_processed = True
        good_price_obj.save()

    GoodPriceType.objects.filter(is_processed=False).delete()
    GoodPriceType.objects.filter(is_processed=True).update(is_processed=False)

    return 'Ok'


def agent_1c_goods(data):

    from application.agent.models import Good, Category, Brand

    Good.objects.all().update(is_processed=False)

    for good in data['goods']:

        try:
            good_obj = Good.objects.get(code=good['code'])
        except Good.DoesNotExist:
            good_obj = Good(code=good['code'])

        # Rest count
        try:
            good_obj.rest = round(good['rest'])
        except:
            good_obj.rest = None

        # Fields
        good_obj.name = good['name'].replace('*', 'x')
        good_obj.unit = good['unit']
        good_obj.price = round(good['price'], 2)
        good_obj.box_count = good['embeddability']
        good_obj.min_count = good['min']
        good_obj.is_processed = True
        good_obj.cashback = good['cashback']
        good_obj.manufacturer = good['manufacturer']
        good_obj.brand_name = good['brend']

        # Brand
        brand_obj = Brand.objects.filter(name__iexact=good['group'].lower()).first()
        if not brand_obj:
            brand_obj = Brand(name=good['group'])
            brand_obj.save()
        good_obj.brand = brand_obj

        # Save
        good_obj.save()

        # # Price
        # try:
        #     gp = GoodPrice.objects.get(good=good_obj, city=def_city)
        #     gp.price = good_obj.price
        #     gp.save()
        # except GoodPrice.DoesNotExist:
        #     GoodPrice(good=good_obj, city=def_city, price=good_obj.price).save()

        # Categories
        groups = []
        for i in data['goods']:
            if i['code'] == good['code']:
                group_obj = Category.objects.filter(name__iexact=i['groupdop'].lower()).first()
                if not group_obj:
                    group_obj = Category(name=i['groupdop'])
                    group_obj.save()
                groups.append(group_obj)
        for i in groups:
            if i not in good_obj.categories.all():
                good_obj.categories.add(i)
        for i in good_obj.categories.all():
            if i not in groups:
                good_obj.categories.remove(i)

    Good.objects.filter(is_processed=False, is_not_delete=False).update(is_public=False)
    Good.objects.filter(is_processed=True).update(is_public=True)

    return 'Ok'


def agent_1c_orders(data):

    from application.agent.models import Order
    from application.solar_staff_accounts.models import SolarStaffAccount

    for payment in data['payment']:

        try:
            try:
                order_obj = Order.objects.get(id=payment['doc'])
            except:
                raise Order.DoesNotExist
            order_obj.from_1c_firm = payment['firm']
            try:
                ss_account = SolarStaffAccount.objects.get(name__iexact=order_obj.from_1c_firm)
                if ss_account:
                    order_obj.ss_account = ss_account
            except:
                pass

            order_obj.from_1c_sum = payment['sum']
            order_obj.from_1c_pay = payment['pay']
            order_obj.from_1c_status = payment['status']
            order_obj.from_1c_cashback = payment['sumcashback']

            if order_obj.status in [1, 4] and order_obj.from_1c_status == 'a' and order_obj.from_1c_sum <= order_obj.from_1c_pay + 15:
                order_obj.status = 2
            if order_obj.status in [1, 4, 2] and order_obj.from_1c_status == 'a' and order_obj.from_1c_sum > order_obj.from_1c_pay + 15:
                order_obj.status = 1
            if order_obj.from_1c_status == 'x':
                order_obj.status = 4
            if order_obj.from_1c_status == 'b':
                order_obj.status = 5

            if order_obj.store is None or order_obj.store.category is None or order_obj.store.category.name != 'Городской отдел':
                order_obj.cashback_sum = order_obj.from_1c_cashback

            order_obj.save()
        except Order.DoesNotExist:
            pass

    return 'Ok'


def tinkoff_payment(data):

    from application.agent.models import TinkoffPayment

    # Получаем пароль
    terminal_key = data['TerminalKey']
    password = ''
    for i in settings.TINKOFF_TERMINALS:
        if i[0] == terminal_key:
            password = i[1]

    # Проверка хеш кода
    values = {}
    for key, value in data.items():
        if key not in ['Receipt', 'Data', 'Token']:
            values[key] = value

    values['Password'] = password

    s = ''
    for key in sorted(values):
        if key == 'Success':
            s += str(values[key]).lower()
        else:
            s += str(values[key])

    hash_str = hashlib.sha256(s.encode('utf-8')).hexdigest()
    if hash_str != data['Token']:
        return 'Error hash'

    # Добавляем в оплаты
    obj = TinkoffPayment()
    if data.get('TerminalKey'):
        obj.terminal_key = str(data['TerminalKey'])[:1000]
    if data.get('OrderId'):
        obj.order_id = str(data['OrderId'])[:1000]
    try:
        obj.success = data['Success']
    except:
        pass
    if data.get('Status'):
        obj.status = str(data['Status'])[:1000]
    if data.get('PaymentId'):
        obj.payment_id = str(data['PaymentId'])[:1000]
    if data.get('ErrorCode'):
        obj.error_code = str(data['ErrorCode'])[:1000]
    if data.get('Amount'):
        obj.amount = str(data['Amount'])[:1000]
    if data.get('CardId'):
        obj.card_id = str(data['CardId'])[:1000]
    if data.get('Pan'):
        obj.pan = str(data['Pan'])[:1000]
    if data.get('ExpDate'):
        obj.exp_date = str(data['ExpDate'])[:1000]
    obj.save()

    # Записываем в заказы Айсмена
    if data.get('PaymentId'):
        orders = IcemanOrder.objects.filter(online_payment_id=data['PaymentId'])
        for i in orders:
            if data.get('Status'):
                i.online_payment_status = str(data['Status'])[:100]
                i.save(update_fields=['online_payment_status'])
                transaction.commit()

    return 'OK'
