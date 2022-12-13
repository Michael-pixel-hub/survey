import copy
import datetime
import hashlib
import math
import openpyxl
import os
import pandas as pd
import random
import re
import subprocess
import traceback

from application.agent.dadata import Dadata
from django.conf import settings
from django.core.cache import cache
from django.core.files import File
from django.db import transaction
from django.db.models import OuterRef, Subquery
from django.utils.translation import ugettext_lazy as _
from preferences.utils import get_setting

from math import sin, cos, sqrt, radians, asin


def get_coordinates(address):

    # Sputnik

    # try:
    #     result = requests.get('http://search.maps.sputnik.ru/search/addr?q=%s' % address, timeout=3)
    #     data_address = result.json()
    #     address_json_data = data_address['result']['address'][0]['features'][0]['geometry']['geometries'][0]['coordinates']
    #     print(address, address_json_data)
    #     return address_json_data[0], address_json_data[1]
    # except:
    #     return None, None

    # Google

    # try:
    #     result = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address=%s&key=AIzaSyCAPzfHUpN1FTqVbuv8vIvmWnv40t94-PU' % address, timeout=5)
    #     data_address = result.json()
    #     address_json_data = data_address['results'][0]['geometry']['location']
    #     return address_json_data['lng'], address_json_data['lat']
    # except:
    #     return None, None

    api_key = get_setting('agent_dadataapikey')
    secret_key = get_setting('agent_dadatasecretkey')

    if not api_key:
        return None, None

    if not secret_key:
        return None, None

    try:
        dadata = Dadata(api_key, secret_key)
        data = dadata.clean('address', address)
        return data['geo_lon'], data['geo_lat']
    except:
        return None, None


# def get_coordinates(address):
#
#     # result = requests.get('https://geocode-maps.yandex.ru/1.x/?format=json&geocode=%s' % address)
#     # data_address = result.json()
#     #
#     # try:
#     #     address_json_data = data_address['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']
#     #     coord = address_json_data['Point']['pos']
#     #     coord = coord.split(' ')
#     #     longitude = coord[0]
#     #     latitude = coord[1]
#     #     return longitude, latitude
#     # except IndexError:
#     #     return None, None
#
#     # #time.sleep(1)
#     #
#     # user_agent = {'User-Agent': 'Free Telegram bot @Serveyor_bot for for merchandising. Mail: mail@engine2.ru'}
#     # result = requests.get('https://nominatim.openstreetmap.org/search?q=%s&format=json' % address, headers=user_agent)
#     # data_address = result.json()
#     #
#     # try:
#     #     address_json_data = data_address[0]
#     #     longitude = address_json_data['lon']
#     #     latitude = address_json_data['lat']
#     #     return longitude, latitude
#     # except IndexError:
#     #     return None, None
#
#     try:
#         result = requests.get('http://search.maps.sputnik.ru/search/addr?q=%s' % address)
#         data_address = result.json()
#         address_json_data = data_address['result']['address'][0]['features'][0]['geometry']['geometries'][0]['coordinates']
#         return address_json_data[0], address_json_data[1]
#     except:
#         return None, None


def import_data_from_file(file, file_name, delete_assortment=False, assign_assortment='store', user=None):

    """
    2,4 - Вторник и четверг
    *2 - 2 раза в неделю
    * - один раз
    **1 - Один раз в месяц
    """

    from .models import Client, Region, Store, Good, Assortment, Import, Category, StoreTask, Task
    from application.survey.models import User

    import_obj = Import(file_name=file_name, user_id=user)
    import_obj.save()

    with open(file, 'rb') as fi:
        import_obj.file = File(fi, name=os.path.basename(fi.name))
        import_obj.save()

    try:

        #rb = xlrd.open_workbook(file)
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb.active
        import_obj.rows_count = sheet.max_row
        import_obj.save()

        factory_code_index = -1

        for row in sheet.iter_rows(values_only=True):
            #row = sheet.row_values(0)
            row = [item.lower() if item is not None else '' for item in row]

            code_index = row.index('код клиента')
            region_index = row.index('регион')
            client_index = row.index('клиент')
            address_index = row.index('адрес')
            good_index = row.index('товар')
            good_code_index = row.index('товар ид')
            group_index = row.index('группа')
            latitude_index = row.index('широта')
            longitude_index = row.index('долгота')
            try:
                factory_code_index = row.index('код завода')
            except ValueError:
                ...
            break

        # StoreTask
        tasks = []
        # tasks_idx = []
        # positions_idx = []
        idx = 0
        # current_task_idx = None
        current_task = None

        for r in row:
            if r == 'задание':
                current_task = {
                    'task_idx': idx, 'position_idx': None, 'period_idx': None, 'hours_idx': None,
                    'user_email_idx': None, 'telegram_idx': None
                }
                tasks.append(current_task)
            if r == 'последовательность' and current_task is not None:
                current_task['position_idx'] = idx
            if r == 'периодичность' and current_task is not None:
                current_task['period_idx'] = idx
            if r == 'время' and current_task is not None:
                current_task['hours_idx'] = idx
            if r == 'время' and current_task is not None:
                current_task['hours_idx'] = idx
            if (r == 'email сюрвеера' or r == 'email') and current_task is not None:
                current_task['user_email_idx'] = idx
            if (r == 'telegram канал' or r == 'telegram') and current_task is not None:
                current_task['telegram_idx'] = idx

            idx += 1


        # for r in row:
        #     if r == 'задание':
        #         tasks_idx.append(idx)
        #         current_task_idx = idx
        #     if r == 'последовательность' and current_task_idx is not None:
        #         positions_idx.append({'task_idx': current_task_idx, 'position_idx': idx})
        #     idx += 1

        #with transaction.atomic():

        # Delete assortment
        if delete_assortment:
            for row in sheet.iter_rows(values_only=True):
                code = row[code_index] if code_index > -1 else ''
                code = code if code is not None else ''
                code = str(code).strip()[:100]
                if not code:
                    continue
                try:
                    store_obj = Store.objects.get(code=code)
                    Assortment.objects.filter(store=store_obj).delete()
                except:
                    pass

        # Process rows
        index = 0
        for row in sheet.iter_rows(values_only=True):

        #for row_num in range(1, sheet.max_row):

            if cache.get('chl_bot_import_cancel'):
                cache.delete('chl_bot_import_cancel')
                import_obj.status = 3
                import_obj.save()
                return

            index += 1
            if index == 1:
                continue

            #row = sheet.row_values(row_num)

            row = list(row)

            s_i = 0
            for s in row:
                if s is None:
                    row[s_i] = ''
                s_i += 1

            code = row[code_index] if code_index > -1 else ''
            factory_code = row[factory_code_index] if factory_code_index > -1 else None
            region = row[region_index] if region_index > -1 else ''
            client = row[client_index] if client_index > -1 else ''
            address = row[address_index] if address_index > -1 else ''
            good = row[good_index] if good_index > -1 else ''
            good_code = row[good_code_index] if good_code_index > -1 else ''
            good_code = str(good_code).replace('.0', '').replace(',0', '')
            group = row[group_index] if group_index > -1 else ''

            try:
                latitude = row[latitude_index].strip() if latitude_index > -1 else ''
            except:
                latitude = row[latitude_index] if latitude_index > -1 else ''
            try:
                longitude = row[longitude_index].strip() if longitude_index > -1 else ''
            except:
                longitude = row[longitude_index] if longitude_index > -1 else ''

            code = str(code).replace('.0', '').strip()[:100]
            region = str(region).strip()[:100]
            client = str(client).strip()[:100]
            address = str(address).strip()[:500]
            good = str(good).strip()[:100]
            group = str(group).strip()[:100]

            region = re.sub(' +', ' ', region).strip()
            client = re.sub(' +', ' ', client).strip()
            address = re.sub(' +', ' ', address).strip()
            good = re.sub(' +', ' ', good).strip()
            group = re.sub(' +', ' ', group).strip()

            if factory_code is not None:
                factory_code = str(factory_code).replace('.0', '').strip()[:1000]
                factory_code = re.sub(' +', ' ', factory_code).strip()

            if not client or not code:
                continue

            client_obj, created = Client.objects.get_or_create(name=client)

            if region:
                region_obj, created = Region.objects.get_or_create(name=region)
            else:
                region_obj = None

            if group:
                group_obj, created = Category.objects.get_or_create(name=group)
            else:
                group_obj = None

            try:
                store_obj = Store.objects.get(code=code)
                store_obj.client = client_obj
                store_obj.category = group_obj
                store_obj.region_o = region_obj
                if store_obj.address != address:
                    store_obj.address = address
                    store_obj.auto_coord = True
                if (store_obj.latitude != latitude or store_obj.longitude != longitude) and latitude and longitude:
                    store_obj.longitude = longitude
                    store_obj.latitude = latitude
                    store_obj.auto_coord = False
                if not store_obj.latitude or not store_obj.longitude:
                    store_obj.auto_coord = True
                if factory_code is not None:
                    store_obj.factory_code = factory_code
                store_obj.save()
            except Store.DoesNotExist:
                store_obj = Store()
                store_obj.code = code
                store_obj.client = client_obj
                store_obj.category = group_obj
                store_obj.region_o = region_obj
                store_obj.address = address
                store_obj.auto_coord = True
                if factory_code is not None:
                    store_obj.factory_code = factory_code
                if latitude and longitude:
                    store_obj.latitude = latitude
                    store_obj.longitude = longitude
                    store_obj.auto_coord = False
                store_obj.save()

            if good and assign_assortment == 'store':
                good_obj, created = Good.objects.get_or_create(name=good)
                if good_code:
                    good_obj.code = good_code
                    good_obj.save()
                Assortment.objects.get_or_create(good=good_obj, store=store_obj)

            for task_data in tasks:

                # Задача
                idx = task_data['task_idx']
                task_name = row[idx]
                task_name = str(task_name).strip()[:100]
                task_name = re.sub(' +', ' ', task_name).strip()

                try:
                    task = Task.objects.get(name=task_name)
                except Task.DoesNotExist:
                    continue

                if good and assign_assortment == 'task':
                    good_obj, created = Good.objects.get_or_create(name=good)
                    if good_code:
                        good_obj.code = good_code
                        good_obj.save()
                    Assortment.objects.get_or_create(good=good_obj, store=store_obj, task=task)

                # Периодичность
                period = None
                if task_data['period_idx'] is not None:
                    idx = task_data['period_idx']
                    period = re.sub(' +', ' ', str(row[idx])).strip()

                if period:

                    st, created = StoreTask.objects.get_or_create(task=task, store=store_obj)
                    if period == '*':
                        st.is_once = True
                        st.per_week = 0
                        st.per_month = None
                        st.days_of_week = ''
                    elif period.startswith('**'):
                        try:
                            months = int(period.replace('**', ''))
                            st.per_month = months
                            st.days_of_week = ''
                            st.per_week = 0
                            st.is_once = False
                        except:
                            pass
                    elif period.startswith('*'):
                        try:
                            days = int(period.replace('*', ''))
                            st.per_week = days
                            st.days_of_week = ''
                            st.is_once = False
                            st.per_month = None
                        except:
                            pass
                    else:
                        st.days_of_week = period.replace('.', ',')
                        st.days_of_week = st.days_of_week.replace(',0', '')
                        st.is_once = False
                        st.per_week = 0
                        st.per_month = None

                    hours = None
                    if task_data['hours_idx'] is not None:
                        idx = task_data['hours_idx']
                        hours = re.sub(' +', ' ', str(row[idx])).strip()

                    if hours and hours != 'None':
                        hours_a = hours.split('-')
                        st.hours_start = int(hours_a[0])
                        st.hours_end = int(hours_a[1])
                    else:
                        st.hours_start = None
                        st.hours_end = None

                    user_email = None
                    if task_data['user_email_idx'] is not None:
                        idx = task_data['user_email_idx']
                        user_email = re.sub(' +', ' ', str(row[idx])).strip()

                    user_obj = None
                    if user_email:
                        try:
                            user_obj = User.objects.filter(email=user_email).first()
                        except:
                            pass
                    st.only_user = user_obj

                    telegram_channel = None
                    if task_data['telegram_idx'] is not None:
                        idx = task_data['telegram_idx']
                        telegram_channel = re.sub(' +', ' ', str(row[idx])).strip().replace('.0', '').replace(',0', '')
                    st.telegram_channel_id = ''
                    if telegram_channel:
                        st.telegram_channel_id = telegram_channel

                    position = None
                    if task_data['position_idx'] is not None:
                        idx = task_data['position_idx']
                        position = re.sub(' +', ' ', str(row[idx])).strip().replace('.0', '').replace(',0', '')
                    if position:
                        st.position = position

                    st.save()

                else:
                    StoreTask.objects.filter(task=task, store=store_obj).delete()

            try:
                Import.objects.get(status=1)
            except Import.DoesNotExist:
                return

            import_obj.rows_process = index
            import_obj.save()

        import_obj.status = 2
        import_obj.date_end = datetime.datetime.now()
        import_obj.save()

    except Exception as e:

        import_obj.status = 4
        # import_obj.report_text = '%s!\n\n%s' % (str(_('Error')), str(e))
        var = traceback.format_exc()
        import_obj.report_text = var
        import_obj.save()
        return


def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    # haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    km = 6367 * c
    return km


def get_duplicate_images(image):

    media_path = os.path.realpath(settings.MEDIA_ROOT)
    image_path = os.path.join(media_path, str(image))
    md5sum = hashlib.md5(open(image_path, 'rb').read()).hexdigest()

    image_path = os.path.join(os.path.realpath(settings.MEDIA_ROOT), os.path.dirname(os.path.abspath(image)))

    output = subprocess.run(
        'find %s -type f -exec md5sum {} + | grep \'^%s\'' % (image_path, md5sum),
        shell=True,
        stdout=subprocess.PIPE,
        universal_newlines=True
    )

    images = []

    s = str(output.stdout).split('\n')
    for i in s:
        sl = i.split('  ')
        try:
            image_path = sl[1]
            image_path = image_path.replace(media_path, '')[1:]
            images.append(image_path)
        except IndexError:
            continue

    return images


def send_user_status_change(te_obj, old_status):

    from application.telegram.models import String
    from application.telegram.tasks import send_message

    if te_obj.source == 'telegram':

        comment = 'Нет'
        if te_obj.comments_status:
            comment = te_obj.comments_status

        s = String.get_string('message_task_execution_status_changed').format(
            task=te_obj.task,
            store=te_obj.store,
            old_status=old_status,
            status=te_obj.get_status_display(),
            comment=comment
        )
        send_message.delay(te_obj.user.telegram_id, s)
    else:
        if te_obj.application == 'shop_survey':
            from application.mobile.models import Notification
        else:
            from application.iceman.models import Notification

        notification = Notification()
        notification.user = te_obj.user
        notification.title = f'Изменение статуса в задаче'
        notification.message = f'В задаче #{te_obj.id} сменился статус с «{old_status}» на ' \
                               f'«{te_obj.get_status_display()}».'
        comment = ''
        if te_obj.comments_status:
            comment = f' Комментарий: {te_obj.comments_status[:1000]}'
        notification.message = f'{notification.message}{comment}'
        notification.category = 'task'
        notification.save()


def get_send_message_data(file):

    from application.survey.models import User

    try:
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb.active

        column_num = None
        column_type = None
        start_row = 1

        for row in sheet.iter_rows():
            cells = list(row)
            column_num = 1
            for cell in cells:
                if str(cell.value).lower() == 'id' or str(cell.value).lower() == 'ид':
                    column_type = 'id'
                    break
                if str(cell.value).lower() == 'e-mail' or str(cell.value).lower() == 'email':
                    column_type = 'email'
                    break
                if str(cell.value).lower() == 'telegram_id':
                    column_type = 'telegram_id'
                    break
                if str(cell.value).lower() == 'telegram login':
                    column_type = 'username'
                    break
                column_num += 1
            if column_type:
                break
            start_row += 1

        if column_type is None:
            return str(_('Error. Excel file data not found'))

        count = 0
        data = []
        row_num = 0
        for row in sheet.iter_rows():

            row_num += 1

            if row_num < start_row + 1:
                continue

            cells = list(row)
            obj = None
            try:
                val = str(cells[column_num-1].value)
                if val:
                    try:
                        obj = User.objects.filter(**{column_type: val}).first()
                    except:
                        pass
            except:
                pass

            if obj:
                count += 1

            fields = []
            fields.append(cells[column_num-1])
            cell_column_num = 0
            for cell in cells:
                cell_column_num += 1
                try:
                    if sheet.cell(start_row, cell_column_num).value == 'Field':
                        fields.append(cell)
                except:
                    pass

            data.append({'obj': obj, 'data': fields})

        return data

        # count = 0
        # for row_num in range(start_row + 1, rows_count):
        #     row = sheet.row_values(row_num)
        #     val = str(row[column_num]).replace('.0', '')
        #     try:
        #         obj = User.objects.filter(**{column_type: val}).first()
        #     except:
        #         obj = None
        #     if obj:
        #         count += 1
        #     data.append({'obj': obj, 'data': row})
        # return data

    except:
        return str(_('Error. Excel file error'))


def get_send_message_data_stores(file):

    from application.agent.models import Store, PromoCode

    promo_codes = PromoCode.objects.filter(is_used=True, store=OuterRef('pk')).values('code')[:1]

    try:

        wb = openpyxl.load_workbook(filename=file)
        sheet = wb.active

        column_num = None
        column_type = None
        start_row = 1

        for row in sheet.iter_rows():
            cells = list(row)
            column_num = 1
            for cell in cells:
                if str(cell.value).lower() == 'id' or str(cell.value).lower() == 'ид':
                    column_type = 'id'
                    break
                column_num += 1
            if column_type:
                break
            start_row += 1

        if column_type is None:
            return str(_('Error. Excel file data not found'))

        count = 0
        data = []
        row_num = 0
        for row in sheet.iter_rows():

            row_num += 1

            if row_num < start_row + 1:
                continue

            cells = list(row)
            obj = None
            try:
                val = str(cells[column_num-1].value)
                if val:
                    try:
                        obj = Store.objects.filter(**{column_type: val}).\
                            annotate(promo_code=Subquery(promo_codes)).first()
                    except:
                        pass
            except:
                pass

            if obj:
                count += 1

            fields = []
            cell_column_num = 0
            for cell in cells:
                cell_column_num += 1
                try:
                    if sheet.cell(start_row, cell_column_num).value == 'Field':
                        fields.append(cell)
                except:
                    pass

            data.append({'obj': obj, 'data': fields})

        return data

    except:
        # import traceback
        # print(traceback.format_exc())
        return str(_('Error. Excel file error'))


def survey_1c_acts(data):

    from application.survey.models import Act, User

    for act in data['act']:

        if not act['id']:
            continue

        try:
            act_obj = Act.objects.get(id_1c=act['id'])
        except:
            act_obj = Act(id_1c=act['id'])

        user_obj = User.objects.filter(email__iexact=act['pers']['email']).first()
        act_obj.user = user_obj

        act_obj.number = act['nomber']
        act_obj.date = act['date']
        act_obj.user_fio = act['pers']['fio']
        act_obj.user_inn = act['pers']['inn']
        act_obj.user_phone = act['pers']['phone']
        act_obj.user_email = act['pers']['email']
        act_obj.sum = act['sum']
        act_obj.date_start = act['start']
        act_obj.date_end = act['end']
        act_obj.save()

    return 'Ok'


def survey_1c_specification(data):

    from application.survey.models import Assortment, Store, Task, Good

    # Получаем задачи
    # tasks_ids = Task.objects.filter(assortment_1c=True).values_list('id', flat=True)

    # Обработка
    clear_stores = []

    for item in data['specification']:

        # Данные
        code_client = item['code_client']
        tovar = str(item['tovar']).strip()

        # Преобразования
        if tovar == 'ДИКСИ Российское 12% в молоч.шокол 80гр*28' or tovar == \
                'Эскимо Российское пломбир в молоч.шокол 80гр*28 ДИКСИ':
            tovar = 'Эскимо Российское пломб/молоч.шок.80гр*40'

        # Ищем магазин
        try:
            store = Store.objects.get(code=code_client)
        except Store.DoesNotExist:
            continue

        # Очищаем
        if code_client not in clear_stores:
            clear_stores.append(code_client)
            Assortment.objects.filter(store=store, task__isnull=True).update(is_delete=True)

        # Ищем товар
        if tovar == '':
            continue
        try:
            good = Good.objects.get(name=tovar)
        except Good.DoesNotExist:
            good = Good(name=tovar, description='')
            good.save()

        # Применяем изменения для задач
        try:
            as_obj = Assortment.objects.filter(task__isnull=True).get(store=store, good=good)
            as_obj.is_delete = False
            as_obj.save()
        except Assortment.DoesNotExist:
            as_obj = Assortment(good=good, store=store)
            as_obj.is_delete = False
            as_obj.save()

    Assortment.objects.filter(is_delete=True).delete()

    return 'Ok'


def excel_get_string(val, max_length=None):

    if type(val) == float:
        val = str(val).replace('.0', '')

    if type(val) == int:
        val = str(val).replace('.', '')

    val = str(val)

    if max_length is not None:
        val = val[:max_length]

    if val == 'nan':
        return None

    return str(val)


@transaction.atomic
def survey_import_tasks_from_file(file, file_name, user=None):

    def save_error(s):
        import_obj.status = 4
        import_obj.report_text = s
        import_obj.save()

    from application.survey.models import Task, User, TasksExecution, ImportTask, Region, Store

    import_obj = ImportTask(file_name=file_name, user_id=user)
    import_obj.save()

    with open(file, 'rb') as fi:
        import_obj.file = File(fi, name=os.path.basename(fi.name))
        import_obj.save()

    required_columns = ['E-mail', 'Сумма', 'Сумма номинал', 'Стоимость задачи', 'Задача', 'Регион', 'Дата начала',
                        'Дата завершения']

    try:

        # Начало
        data = pd.read_excel(file, sheet_name=0, header=0)
        import_obj.rows_count = len(data)
        import_obj.save()

        # Обязательные колонки
        for i in required_columns:
            if i not in data.columns:
                save_error(f'Колонка "{i}" не обнаружена в Excel-файле. '
                           f'Названия колонок должно быть в первой строке таблицы.')
                return

        # Обработка
        sid = transaction.savepoint()
        for index, row in data.iterrows():

            # Отмена
            if cache.get('survey_import_tasks_cancel'):
                transaction.savepoint_rollback(sid)
                cache.delete('survey_import_tasks_cancel')
                import_obj.status = 3
                import_obj.save()
                return

            # Пользователь
            user_mail = excel_get_string(row['E-mail'], 100)
            user = User.objects.filter(email=user_mail)
            if user.count() > 1:
                transaction.savepoint_rollback(sid)
                save_error(f'Существует несколько пользователей с почтой "{user_mail}".')
                return
            user = user.first()
            if user is None:
                transaction.savepoint_rollback(sid)
                save_error(f'Пользователь с почтой "{user_mail}" не найден.')
                return

            # Сумма
            summa = row['Сумма']
            try:
                if type(summa) == str:
                    summa = float(summa.replace(',', '.'))
                summa = round(summa, 2)
            except:
                transaction.savepoint_rollback(sid)
                save_error(f'Сумма "{summa}" не является числом.')
                return

            # Сумма
            summa_source = row['Сумма номинал']
            try:
                if type(summa_source) == str:
                    summa_source = float(summa_source.replace(',', '.'))
                summa_source = round(summa_source, 2)
            except:
                transaction.savepoint_rollback(sid)
                save_error(f'Сумма номинал "{summa_source}" не является числом.')
                return

            # Сумма
            price = row['Стоимость задачи']
            try:
                if type(price) == str:
                    price = float(price.replace(',', '.'))
                price = round(price, 2)
            except:
                transaction.savepoint_rollback(sid)
                save_error(f'Стоимость задачи "{price}" не является числом.')
                return

            # Задача
            task_name = excel_get_string(row['Задача'], 100)
            try:
                task = Task.objects.get(name=task_name)
            except Task.DoesNotExist:
                transaction.savepoint_rollback(sid)
                save_error(f'Задача "{task_name}" не найдена.')
                return

            # Регион
            region_name = excel_get_string(row['Регион'], 100)
            try:
                region = Region.objects.get(name=region_name)
            except Region.DoesNotExist:
                transaction.savepoint_rollback(sid)
                save_error(f'Задача "{region_name}" не найдена.')
                return

            # Дата начала
            date_start = row['Дата начала']
            try:
                date_start = date_start.to_pydatetime()
            except:
                transaction.savepoint_rollback(sid)
                save_error(f'Дата "{date_start}" должна быть формата дд.мм.ГГГГ')
                return

            # Дата завершения
            date_end = row['Дата завершения']
            try:
                date_end = date_end.to_pydatetime()
            except:
                transaction.savepoint_rollback(sid)
                save_error(f'Дата "{date_end}" должна быть формата дд.мм.ГГГГ')
                return

            # Комментарий
            comment = excel_get_string(row['Комментарий'])

            # Строим массив временных меток
            dates = []
            days = (date_end - date_start).days + 1
            for i in range(0, days):
                date = date_start + datetime.timedelta(days=i)
                for h in range(8, 18):
                    for m in (0, 15, 30, 45):
                        minutes = m + random.randint(0, 10)
                        dates.append(
                            date +
                            datetime.timedelta(hours=h) +
                            datetime.timedelta(minutes=minutes) +
                            datetime.timedelta(seconds=random.randint(0, 59))
                        )

            # Считаем количество задач
            tasks_count = math.ceil(float(summa) / float(price))

            # Кол-во задач больше, чем количество дат
            if tasks_count > len(dates):
                transaction.savepoint_rollback(sid)
                save_error(f'Невозможно создать такое количество задач в этот промежуток времени. '
                           f'Пользователь может выполнить максимум 40 задач в сутки. '
                           f'Вы можете увелечить стоимость задания или промежуток времени.')
                return

            # Создаем задачи
            current_summa = 0
            current_summa_source = 0
            for i in range(0, tasks_count):

                # Дата
                date_index = random.randint(0, len(dates) - 1)
                start_date = dates[date_index]
                del(dates[date_index])

                # Создаем задачу
                te = TasksExecution()
                te.user = user
                te.task = task
                te.date_start = start_date
                te.date_end = te.date_start + datetime.timedelta(seconds=random.randint(120, 240))
                te.date_end_user = te.date_end
                te.status = 3
                te.store = Store.objects.filter(region_o=region).order_by('?').first()
                te.comments = '' if comment is None else comment
                te.source = 'admin'
                te.application = task.application

                # Магазин
                if te.store is None:
                    transaction.savepoint_rollback(sid)
                    save_error(f'В регионе {region_name} нет магазинов.')
                    return

                te.longitude = te.store.longitude
                te.latitude = te.store.latitude

                # Стоимость
                if i != tasks_count - 1:
                    te.money = round(price, 2)
                else:
                    te.money = round(summa - current_summa, 2)
                if i != tasks_count - 1:
                    te.money_source = round(summa_source / summa * te.money, 2)
                else:
                    te.money_source = round(summa_source - current_summa_source, 2)

                # Добавляем
                current_summa += te.money
                current_summa_source += te.money_source
                te.save()
                te.date_start = start_date
                te.save(update_fields=['date_start'])

        # Завершение
        transaction.savepoint_commit(sid)
        import_obj.status = 2
        import_obj.date_end = datetime.datetime.now()
        import_obj.save()

    except Exception as e:
        save_error(e)
        return


def survey_1c_self_employers(data):

    from application.survey.models import User

    for item in data['self_employers']:

        email = item['email']
        status = item['status']

        users = User.objects.filter(email=email)

        for user in users:
            user.status_legal = 'self_employed' if status else 'other'
            user.save(update_fields=['status_legal'])

    return 'Ok'
