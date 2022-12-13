import datetime
import io
import openpyxl
import os
import random
import re
import shutil
import time
import traceback
import xlwt

from application.agent.models import Order as AgentOrder, OrderGood as AgentOrderGood
from application.loyalty.models import Department, Program
from collections import Counter
from dal import autocomplete
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.messages import get_messages
from django.core.cache import cache
from django.db.models import Count, Q, Value
from django.db.models.functions import Concat
from django.db.models import (OuterRef, Subquery, Count, Min, Max, FloatField, DateField, Exists,
                              ExpressionWrapper, IntegerField, DateTimeField, CharField)
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.translation import ugettext_lazy as _
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, TemplateView, View
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.writer.excel import save_virtual_workbook
from PIL import Image, ImageFont, ImageDraw
from xlsxwriter.workbook import Workbook
from django.core.paginator import Paginator

from .forms import ImportForm, ExportForm, ImportTasksForm, UploadImageForm
from .models import Import, TasksExecution, Client, Store, Region, StoreTask, Task, Category, UploadRequests, User, \
    TasksExecutionImage, Rank, Act, ExternalRequests, ImportTask, UserStatus, UserSub, TaskStep, UserDevice
from .tasks import import_data, survey_import_tasks

from application.supervisor.models import UserSchedule, UserScheduleTaskExecution
from application.users.models import User as AuthUser


def find_weeks(start_date, end_date):
    # subtract_days = start_date.isocalendar()[2] - 1
    # current_date = start_date + datetime.timedelta(days=7-subtract_days)
    current_date = start_date
    weeks_between = []
    while current_date <= end_date:
        weeks_between.append(
            '{}{:02d}'.format(*current_date.isocalendar()[:2])
        )
        current_date += datetime.timedelta(days=7)
    return weeks_between


def dates_between(start, end):
    while start <= end:
        yield start
        start += datetime.timedelta(1)


def count_weekday(start, end):
    counter = Counter()
    for date in dates_between(start, end):
        counter[date.strftime('%a')] += 1
    return counter


class Upload1cView(View):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        super(Upload1cView, self).dispatch(request, *args, **kwargs)
        r = UploadRequests()
        r.request_text = request.body.decode('utf-8')
        try:
            r.request_ip = request.META['REMOTE_ADDR']
        except:
            pass
        r.request_method = request.method
        r.request_type = request.content_type
        r.request_files = ''
        for i in request.FILES.values():
            r.request_files += str(i.name) + '\n'
        r.save()
        return HttpResponse('Ok')


class ExternalRequestView(View):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        super().dispatch(request, *args, **kwargs)
        r = ExternalRequests()
        r.request_text = request.body.decode('utf-8')
        try:
            r.request_ip = request.META['REMOTE_ADDR']
        except:
            pass
        r.request_method = request.method
        r.request_type = request.content_type
        r.request_files = ''
        for i in request.FILES.values():
            r.request_files += str(i.name) + '\n'
        r.save()
        return HttpResponse('OK')


class OSMMapView(TemplateView):
    template_name = 'survey/map_osm.html'

    def dispatch(self, request, *args, **kwargs):
        return super(OSMMapView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Context
        context = super(OSMMapView, self).get_context_data(**kwargs)

        location = self.request.GET.get('location')
        if not location:
            location = '55.751407, 37.618877'
        context['location'] = location

        # Exit
        return context


class SmorozaMapView(TemplateView):
    template_name = 'survey/map_smoroza.html'

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Context
        context = super().get_context_data(**kwargs)

        # Exit
        return context


class MapView(TemplateView):
    template_name = 'survey/map.html'

    def dispatch(self, request, *args, **kwargs):
        return super(MapView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Context
        context = super(MapView, self).get_context_data(**kwargs)

        location = self.request.GET.get('location')
        if not location:
            location = '55.751407, 37.618877'
        context['location'] = location

        # Search
        tasks = Task.objects.filter(is_public=True)
        context['tasks'] = tasks
        context['rng'] = random.uniform(1.0, 1.9)

        # Exit
        return context


class ImportView(FormView):
    template_name = 'survey/import.html'
    form_class = ImportForm
    success_url = reverse_lazy('survey:import')
    is_success = False

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        if self.request.GET.get('action') == 'cancel':
            cache.set('survey_bot_import_cancel', 1)
        return super(ImportView, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):

        import hashlib
        import random

        try:
            Import.objects.get(status=1)
            return
        except Import.DoesNotExist:
            pass

        temp_file = '/tmp/%s.xlsx' % hashlib.sha1(str(random.random()).encode()).hexdigest()[:12]

        file_name = self.request.FILES['file'].name

        # # Delete
        # if form.cleaned_data['delete_old']:
        #     Assortment.objects.all().delete()

        with open(temp_file, 'wb+') as destination:
            for chunk in self.request.FILES['file'].chunks():
                destination.write(chunk)
            destination.close()
            import_data.delay(temp_file, file_name, form.cleaned_data['delete_assortment'],
                              form.cleaned_data['assortment_type'], user=self.request.user.id)

        messages.success(self.request, _('Import data task sending successfully'))

        return super(ImportView, self).form_valid(form)

    def get_context_data(self, **kwargs):

        # Context
        context = super(ImportView, self).get_context_data(**kwargs)

        # Search
        try:
            import_obj = Import.objects.get(status=1)
        except Import.DoesNotExist:
            import_obj = None

        # Messages
        storage = get_messages(self.request)
        for i in storage:
            self.is_success = True

        context['import_obj'] = import_obj
        context['is_success'] = self.is_success

        # Last import
        last_import = Import.objects.first()
        context['last_import'] = last_import

        # Exit
        return context


class ReportsView(TemplateView):
    template_name = 'survey/reports.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportsView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Context
        context = super(ReportsView, self).get_context_data(**kwargs)

        # Task execution report
        context['te_report_statuses'] = TasksExecution.statuses
        context['agent_orders_statuses'] = AgentOrder.statuses
        context['regions'] = Region.objects.all()
        context['loyalty_departments'] = Department.objects.filter(is_public=True)
        context['loyalty_programs'] = Program.objects.filter(is_public=True)
        context['taxpayers_users'] = settings.TAXPAYERS_STAFF_USERS

        # Exit
        return context


class ReportSurveyorsView(TemplateView):
    """
    Отчет по сурвеерам (отчет №3)
    """
    template_name = 'survey/reports.surveyors.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportSurveyorsView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        # Context
        context = super(ReportSurveyorsView, self).get_context_data(**kwargs)

        try:
            datetime_start = self.request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = self.request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        try:
            region = self.request.GET.getlist('region')
        except (TypeError, ValueError):
            region = None

        try:
            route = self.request.GET.get('route')
            if route is None:
                route = ''
        except (TypeError, ValueError):
            route = ''

        try:
            task = self.request.GET.getlist('task')
        except (TypeError, ValueError):
            task = None

        try:
            superviser = self.request.GET.get('superviser')
        except (TypeError, ValueError):
            superviser = None

        if not datetime_start:
            datetime_start = datetime.datetime.today()
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        if not datetime_end:
            datetime_end = datetime.datetime.today()
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)

        context['datetime_start'] = datetime_start
        context['datetime_start_d'] = datetime_start.strftime('%Y-%m-%d')
        context['datetime_end'] = datetime_end
        context['datetime_end_d'] = datetime_end.strftime('%Y-%m-%d')
        context['region'] = region
        context['superviser'] = superviser
        context['regions_selected'] = Region.objects.filter(id__in=region)
        context['tasks_selected'] = Task.objects.filter(id__in=task)
        context['superviser_selected'] = User.objects.filter(id=superviser)
        context['route'] = route

        te_count = UserScheduleTaskExecution.objects.filter(Q(task_id__in=task) if task else Q(),
            schedule__date__gte=datetime_start, schedule__date__lte=datetime_end,
            schedule__user_id=OuterRef('user_id')).values('schedule__user_id').annotate(
            count=Count('schedule__user_id')).order_by().values('count')

        te_count_done = UserScheduleTaskExecution.objects.filter(Q(task_id__in=task) if task else Q(),
            te__isnull=False, te__date_start__gte=datetime_start, te__date_end__lte=datetime_end,
            schedule__user_id=OuterRef('user_id')).values('schedule__user_id').annotate(
            count=Count('schedule__user_id')).order_by().values('count')

        te_min_date = UserScheduleTaskExecution.objects.filter(Q(task_id__in=task) if task else Q(),
            te__isnull=False, te__date_start__gte=datetime_start, te__date_end__lte=datetime_end,
            schedule__user_id=OuterRef('user_id')).values('schedule__user_id').annotate(
            min=Min('te__date_start')).order_by().values('min')

        te_max_date = UserScheduleTaskExecution.objects.filter(Q(task_id__in=task) if task else Q(),
            te__isnull=False, te__date_start__gte=datetime_start, te__date_end__lte=datetime_end,
            schedule__user_id=OuterRef('user_id')).values('schedule__user_id').annotate(
            max=Max('te__date_start')).order_by().values('max')

        main_filter = {}
        if datetime_start:
            main_filter['date__gte'] = datetime_start

        if datetime_end:
            main_filter['date__lte'] = datetime_end

        if route or (not route and not region and not superviser):
            main_filter['user__route__icontains'] = route

        schedule = UserSchedule.objects.select_related('user').filter(**main_filter).values(
            'user_id', 'user__name', 'user__surname', 'user__email', 'user__status__name', 'user__route').distinct().annotate(
            te_count=Subquery(te_count, output_field=IntegerField()),
            te_count_done=Subquery(te_count_done, output_field=IntegerField()),
            te_min_date=Subquery(te_min_date, output_field=DateField()),
            te_max_date=Subquery(te_max_date, output_field=DateField()),
        )
        
        route_region_filter = {}
        if region or (not route and not region and not superviser):
            route_region_filter['store__region_o__in'] = region

        superviser_filter = {}
        if superviser:
            superviser_filter['only_user_id__in'] = UserSub.objects.select_related('user_sub').filter(
                user_id=superviser).order_by('user_sub__id').distinct('user_sub__id').values('user_sub__id')

        region_users = StoreTask.objects.filter(
            Q(task__id__in=task) if task else Q(),
            only_user_id=OuterRef('user_id'),
            **route_region_filter,
            **superviser_filter
        )
        schedule = schedule.annotate(region_users=Exists(region_users)).filter(region_users=True)

        schedule = schedule.order_by('user__surname', 'user__name')

        content = []
        row = {}
        if_true = lambda x: x if x else ''
        if_true_zero = lambda x: x if x else 0
        for user in schedule:
            if user['te_count_done'] and user['te_count'] and user['te_count'] > 0:
                progress = int(user['te_count_done']/user['te_count']*100)
            else:
                progress = 0
            if progress < 33:
                progress_color = '#ec9399'
            elif progress >= 33 and progress < 66:
                progress_color = '#fbf894'
            elif progress >= 66:
                progress_color = '#98fc98'
            row = {
                'route': str(if_true(user['user__route'])),
                'user_id': user['user_id'],
                'fio': User(name=user['user__name'], surname=user['user__surname']).fio + ', '\
                           + if_true(user['user__email']) + ', ' + str(if_true(user['user__status__name'])),
                'te_count': str(if_true_zero(user['te_count'])),
                'te_count_done': str(if_true_zero(user['te_count_done'])),
                'te_min_date': user['te_min_date'],
                'te_max_date': user['te_max_date'],
                'progress': progress,
                'progress_color': progress_color
            }
            content.append(row)

        data_r = []
        data_r.append({'data': content})
        context['data'] = data_r

        # Exit
        return context


class ReportTeView(View):
    template_name = 'survey/reports.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportTeView, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        if datetime_start is None or datetime_end is None:
            messages.error(self.request, 'Выберите дату дачала и дату завершения для отчета.')
            return redirect(reverse_lazy('survey:reports'))

        if (datetime_end - datetime_start).days > 31:
            messages.error(self.request, 'Максимальное количество дней отчета: 31')
            return redirect(reverse_lazy('survey:reports'))

        try:
            task = self.request.GET.getlist('task')
        except (TypeError, ValueError):
            task = None

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(_('Task execution report')))

        te = TasksExecution.objects.all().select_related('store', 'task', 'user', 'store__client', 'store__region_o',
                                                         'check_user', 'store_iceman')
        te = te.extra(
            select={
                'questions': "SELECT string_agg(CONCAT(q1.constructor_step_name, '. ', q1.question), ';;' "
                             "order by q1.id) FROM chl_tasks_executions_questionnaires q1 "
                             "where q1.Task_id = chl_tasks_executions.id",
                'answers': "SELECT string_agg(q2.answer, ';;' order by q2.id) "
                           "FROM chl_tasks_executions_questionnaires q2 where q2.Task_id = chl_tasks_executions.id",
            }
        )
        tasks_id = self.request.user.task.values_list('id', flat=True)
        if tasks_id:
            te = te.filter(task__in=tasks_id)
        if task:
            te = te.filter(task__in=task)
        if datetime_start:
            te = te.filter(date_start__gte=datetime_start)
        if datetime_end:
            te = te.filter(date_start__lte=datetime_end)
        if request.GET.get('status'):
            te = te.filter(status=request.GET.get('status'))

        # Excel
        output = io.BytesIO()
        workbook = Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        all_format_dict = {'font_size': 9}
        a_fmt = workbook.add_format(all_format_dict)
        format_dict = {'num_format': 'dd.mm.yyyy hh:mm:ss', 'font_size': 9}
        d_fmt = workbook.add_format(format_dict)

        if not tasks_id:

            headers = ['Дата', 'Дата завершения', 'Ид пользователя', 'Фамилия', 'Имя', 'E-mail', 'Менеджер',
                       'Название маршрута', 'Задача', 'Сумма, руб.', 'Статус', 'Код магазина', 'Код завода', 'Адрес',
                       'Клиент', 'Регион', 'Регион пользователя', 'Комментарии', 'Чек', 'Аудитор',
                       'Проверено аудитором', 'Цена номинал', 'Источник']
            widths = [15, 15, 17, 15, 15, 25, 22, 20, 22, 15, 12, 15, 14, 50, 16, 18, 25, 30, 13, 10, 23, 16, 13]

            # Make Excel data
            row_num = 0
            for i in te:
                row_num += 1
                worksheet.write(row_num, 0, i.date_start, d_fmt)
                worksheet.write(row_num, 1, i.date_end, d_fmt)
                worksheet.write(row_num, 2, i.user.id, a_fmt)
                worksheet.write(row_num, 3, i.user.surname, a_fmt)
                worksheet.write(row_num, 4, i.user.name, a_fmt)
                worksheet.write(row_num, 5, i.user.email if i.user.email is not None else '', a_fmt)
                worksheet.write(row_num, 6, i.user.advisor, a_fmt)
                worksheet.write(row_num, 7, i.user.route, a_fmt)
                worksheet.write(row_num, 8, i.task.name if i.task else '', a_fmt)
                worksheet.write(row_num, 9, i.money, a_fmt)
                worksheet.write(row_num, 10, i.get_status_display(), a_fmt)
                if i.store_iceman is None:
                    worksheet.write(row_num, 11, i.store.code if i.store is not None else '', a_fmt)
                    worksheet.write(row_num, 12, i.store.factory_code if i.store and i.store.factory_code is None else '', a_fmt)
                    worksheet.write(row_num, 13, i.store.address if i.store is not None else '', a_fmt)
                    worksheet.write(row_num, 14, i.store.client.name if i.store is not None else '', a_fmt)
                    worksheet.write(row_num, 15,
                                    i.store.region_o.name
                                    if i.store is not None and i.store.region_o is not None else '',
                                    a_fmt)
                else:
                    worksheet.write(row_num, 11, i.store_iceman.code if i.store_iceman is not None else '', a_fmt)
                    worksheet.write(row_num, 12, '', a_fmt)
                    worksheet.write(row_num, 13, i.store_iceman.address if i.store_iceman is not None else '', a_fmt)
                    worksheet.write(row_num, 14, i.store_iceman.name if i.store_iceman is not None else '', a_fmt)
                    worksheet.write(row_num, 15,
                                    i.store_iceman.region.name
                                    if i.store_iceman is not None and i.store_iceman.region is not None else '',
                                    a_fmt)
                worksheet.write(row_num, 16, i.user.city, a_fmt)
                worksheet.write(row_num, 17, i.comments, a_fmt)
                worksheet.write(row_num, 18, str(i.get_check_type_display()), a_fmt)
                worksheet.write(row_num, 19, 'Да' if i.is_auditor else 'Нет', a_fmt)
                worksheet.write(row_num, 20, str(i.check_user) if i.check_user is not None else '', a_fmt)
                worksheet.write(row_num, 21, i.money_source, a_fmt)
                worksheet.write(row_num, 22, i.get_source_display(), a_fmt)
                curr_column = 23
                curr_index = 0
                if i.questions and i.answers:
                    questions = i.questions.split(';;')
                    answers = i.answers.split(';;')
                    for question in questions:
                        if len(headers) <= curr_column:
                            headers.append('Вопрос анкеты')
                            headers.append('Ответ анкеты')
                            widths.append(18)
                            widths.append(18)
                        worksheet.write(row_num, curr_column, question, a_fmt)
                        ws.col(curr_column).width = 10000
                        ws.col(curr_column + 1).width = 10000
                        try:
                            worksheet.write(row_num, curr_column + 1, answers[curr_index], a_fmt)
                        except:
                            pass
                        curr_index += 1
                        curr_column += 2


        else:

            headers = ['Дата', 'Дата завершения', 'Ид пользователя', 'E-mail', 'Задача', 'Статус', 'Код магазина',
                       'Код завода', 'Адрес', 'Клиент', 'Регион', 'Регион пользователя', 'Комментарии', 'Источник']

            widths = [15, 15, 17, 25, 22, 12, 15, 14, 50, 16, 18, 25, 30, 13]

            # Make Excel data
            row_num = 0
            for i in te:
                row_num += 1
                worksheet.write(row_num, 0, i.date_start, d_fmt)
                worksheet.write(row_num, 1, i.date_end, d_fmt)
                worksheet.write(row_num, 2, i.user.id, a_fmt)
                worksheet.write(row_num, 3, i.user.email, a_fmt)
                worksheet.write(row_num, 4, i.task.name if i.task else '', a_fmt)
                worksheet.write(row_num, 5, i.get_status_display(), a_fmt)
                worksheet.write(row_num, 6, i.store.code if i.store is not None else '', a_fmt)
                worksheet.write(row_num, 7, i.store.factory_code if i.store is not None and i.store.factory_code is not None else '', a_fmt)
                worksheet.write(row_num, 8, i.store.address if i.store is not None else '', a_fmt)
                worksheet.write(row_num, 9, i.store.client.name if i.store is not None else '', a_fmt)
                worksheet.write(row_num, 10, i.store.region_o.name if i.store is not None and i.store.region_o is not None else '', a_fmt)
                worksheet.write(row_num, 11, i.user.city, a_fmt)
                worksheet.write(row_num, 12, i.comments, a_fmt)
                worksheet.write(row_num, 13, i.get_source_display(), a_fmt)
                curr_column = 14
                curr_index = 0
                if i.questions and i.answers:
                    questions = i.questions.split(';;')
                    answers = i.answers.split(';;')
                    for question in questions:
                        if len(headers) <= curr_column:
                            headers.append('Вопрос анкеты')
                            headers.append('Ответ анкеты')
                            widths.append(18)
                            widths.append(18)
                        worksheet.write(row_num, curr_column, question, a_fmt)
                        ws.col(curr_column).width = 10000
                        ws.col(curr_column + 1).width = 10000
                        try:
                            worksheet.write(row_num, curr_column + 1, answers[curr_index], a_fmt)
                        except:
                            pass
                        curr_index += 1
                        curr_column += 2


        header_format_dict = {'bg_color': '#eeeeee', 'bold': True, 'border': 1}
        h_fmt = workbook.add_format(header_format_dict)

        worksheet.write_row(0, 0, headers, h_fmt)
        worksheet.autofilter(0, 0, 1, len(headers) - 1)

        for i in range(len(widths)):
            worksheet.set_column(i, i, width=widths[i])

        # Done
        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="report_task_execution_%s.xls"' % \
                                          datetime.datetime.now()

        output.close()

        return response


class ReportActsView(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        def save_to_cell(ws, row, column, value):

            font = Font(name='Arial', size=8)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            cell = ws.cell(row, column)
            cell.value = value
            cell.font = font
            cell.border = border

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="acts_%s.xls"' % \
                                          datetime.datetime.now()

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Acts'))

        # items
        items = Act.objects.all()
        if datetime_start:
            items = items.filter(date__gte=datetime_start)
        if datetime_end:
            items = items.filter(date__lte=datetime_end)

        # Настройки
        start_row = 5
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        header = (
            ('Номер', 8),
            ('Дата', 13),
            ('1с ид', 27),
            ('Сумма', 8),
            ('Пользователь ид', 16),
            ('Фио', 26),
            ('ИНН', 10),
            ('Телефон', 12),
            ('E-mail', 20),
            ('Дата начала', 13),
            ('Дата завершения', 16),
            ('Дата статуса', 14),
            ('Статус', 12),
            ('Чек', 44),
        )

        # Title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
        cell = ws['A2']
        cell.value = str(_('Acts {date_start} {date_end}').format(
            date_start='с ' + datetime_start.strftime('%d.%m.%Y') if datetime_start else '',
            date_end='по ' + datetime_end.strftime('%d.%m.%Y') if datetime_end else '',
        ))
        cell.font = Font(name='Arial', size=14, bold=True)
        rd = ws.row_dimensions[2]
        rd.height = 18

        # Rows
        row = start_row
        for i in items:
            row += 1

            rd = ws.row_dimensions[row]
            rd.height = 13

            save_to_cell(ws, row, 1, i.number)
            save_to_cell(ws, row, 2, i.date)
            save_to_cell(ws, row, 3, i.id_1c)
            save_to_cell(ws, row, 4, i.sum)
            save_to_cell(ws, row, 5, i.user_id)
            save_to_cell(ws, row, 6, i.user_fio)
            save_to_cell(ws, row, 7, i.user_inn)
            save_to_cell(ws, row, 8, i.user_phone)
            save_to_cell(ws, row, 9, i.user_email)
            save_to_cell(ws, row, 10, i.date_start)
            save_to_cell(ws, row, 11, i.date_end)
            save_to_cell(ws, row, 12, i.date_update)
            save_to_cell(ws, row, 13, i.get_check_type_display())
            save_to_cell(ws, row, 14, i.url)

        # Filters
        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='f4ecc5', end_color='f4ecc5', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), items.count() + start_row)
        ws.auto_filter.ref = filter_range

        response.content = save_virtual_workbook(wb)
        return response


class ReportTaxpayersView(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        if request.user.email not in settings.TAXPAYERS_STAFF_USERS:
            return HttpResponseForbidden()
        return super(ReportTaxpayersView, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        def save_to_cell(ws, row, column, value):

            font = Font(name='Arial', size=8)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            cell = ws.cell(row, column)
            cell.value = value
            cell.font = font
            cell.border = border

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report_orders_%s.xls"' % \
                                          datetime.datetime.now()

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Taxpayers users'))

        # items
        items = User.objects.filter(taxpayer_status=True).select_related('taxpayer_bank')
        if datetime_start:
            items = items.filter(taxpayer_date__gte=datetime_start)
        if datetime_end:
            items = items.filter(taxpayer_date__lte=datetime_end)

        # Настройки
        start_row = 1
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        header = (
            ('Фамилия', 30),
            ('Имя', 30),
            ('Отчество', 30),
            ('Паспорт серия', 30),
            ('Паспорт номер', 30),
            ('Инн Самозанятого', 30),
            ('Банк', 12),
            ('Номер счета', 15),
            ('Телефон ', 30),
            ('Email ', 30),
        )

        # # Title
        # ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
        # cell = ws['A2']
        # cell.value = str(_('Taxpayers users {date_start} {date_end}').format(
        #     date_start='с ' + datetime_start.strftime('%d.%m.%Y') if datetime_start else '',
        #     date_end='по ' + datetime_end.strftime('%d.%m.%Y') if datetime_end else '',
        # ))
        # cell.font = Font(name='Arial', size=14, bold=True)
        # rd = ws.row_dimensions[2]
        # rd.height = 18

        # Rows
        row = start_row
        for i in items:
            row += 1

            rd = ws.row_dimensions[row]
            rd.height = 13

            save_to_cell(ws, row, 1, i.taxpayer_surname)
            save_to_cell(ws, row, 2, i.taxpayer_name)
            save_to_cell(ws, row, 3, i.taxpayer_patronymic)
            save_to_cell(ws, row, 4, i.taxpayer_passport_series)
            save_to_cell(ws, row, 5, i.taxpayer_passport_number)
            save_to_cell(ws, row, 6, i.taxpayer_inn)
            save_to_cell(ws, row, 7, i.taxpayer_bank.name if i.taxpayer_bank else '')
            save_to_cell(ws, row, 8, i.taxpayer_bank_account)
            save_to_cell(ws, row, 9, re.sub('\D', '', i.taxpayer_phone))
            save_to_cell(ws, row, 10, i.taxpayer_email)

        # Filters
        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='f4ecc5', end_color='f4ecc5', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), items.count() + start_row)
        ws.auto_filter.ref = filter_range

        response.content = save_virtual_workbook(wb)
        return response


class ReportOrdersView(View):
    template_name = 'survey/reports.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportOrdersView, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report_orders_%s.xls"' % \
                                          datetime.datetime.now()

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Agent bot orders'))

        # Title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        cell = ws['A2']
        cell.value = str(_('Agent bot orders {date_start} {date_end}').format(
            date_start='с ' + datetime_start.strftime('%d.%m.%Y') if datetime_start else '',
            date_end='по ' + datetime_end.strftime('%d.%m.%Y') if datetime_end else '',
        ))
        cell.font = Font(name='Arial', size=14, bold=True)
        rd = ws.row_dimensions[2]
        rd.height = 18

        # Orders
        items = AgentOrderGood.objects.all().prefetch_related('order', 'order__user', 'brand', 'category',
                                                              'order__store', 'order__store__city',
                                                              'order__store__category')
        if datetime_start:
            items = items.filter(order__date_order__gte=datetime_start)
        if datetime_end:
            items = items.filter(order__date_order__lte=datetime_end)
        if request.GET.get('status'):
            items = items.filter(order__status=request.GET.get('status'))

        # Настройки
        start_row = 4
        font = Font(name='Arial', size=8)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Rows
        row = start_row
        for i in items:
            row += 1

            rd = ws.row_dimensions[row]
            rd.height = 13

            cell = ws.cell(row, 1)
            cell.value = i.order.id
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 2)
            cell.value = i.order.date_order
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 3)
            cell.value = i.order.delivery_date
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 4)
            cell.value = i.order.user.email
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 5)
            cell.value = i.order.store.phone if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 6)
            cell.value = i.order.store.contact if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 7)
            cell.value = i.order.user.advisor
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 8)
            cell.value = i.order.store.id if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 9)
            cell.value = i.order.store.name if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 10)
            cell.value = i.order.store.category.name if i.order.store and i.order.store.category else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 11)
            cell.value = i.order.store.agent if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 12)
            cell.value = i.order.store.inn if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 13)
            cell.value = str(i.order.store.city) if i.order.store else ''
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 14)
            cell.value = i.order.delivery_address
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 15)
            cell.value = i.order.comment
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 16)
            cell.value = i.order.get_status_display()
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 17)
            cell.value = i.code
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 18)
            cell.value = i.name
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 19)
            cell.value = str(i.brand)
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 20)
            cell.value = str(i.category)
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 21)
            cell.value = i.unit
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 22)
            cell.value = i.price
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 23)
            cell.value = i.count
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 24)
            cell.value = i.sum
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 25)
            cell.value = i.order.sum
            cell.font = font
            cell.border = border

            cell = ws.cell(row, 26)
            cell.value = i.order.cashback_sum
            cell.font = font
            cell.border = border

        # Header
        header = (
            ('Ид', 6),
            ('Дата заказа', 15),
            ('Дата доставки', 16),
            ('Заказчик', 22),
            ('Телефон', 14),
            ('Контактное лицо', 18),
            ('Рекомендатель', 17),
            ('Id магазина', 14),
            ('Название магазина', 19),
            ('Категория магазина', 19),
            ('Торговый представитель', 21),
            ('ИНН', 13),
            ('Регион', 15),
            ('Адрес доставки', 30),
            ('Комментарий', 30),
            ('Статус', 14),
            ('Код товара', 14),
            ('Товар', 30),
            ('Бренд', 14),
            ('Категория', 14),
            ('Ед. измерения', 16),
            ('Цена', 8),
            ('Количество', 13),
            ('Сумма', 9),
            ('Сумма заказа', 15),
            ('Кешбек', 10),
        )
        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='f4ecc5', end_color='f4ecc5', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), items.count() + start_row)
        ws.auto_filter.ref = filter_range

        response.content = save_virtual_workbook(wb)
        return response


class TaskClientsAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return Client.objects.none()

        qs = Client.objects.all()

        if self.q:
            qs = qs.filter(name__icontains=self.q)

        return qs


class TaskStoresAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return Store.objects.none()

        qs = Store.objects.all()

        if self.q:
            qs = qs.filter(Q(code__icontains=self.q) | Q(address__icontains=self.q))

        return qs


class TaskRegionsAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return Region.objects.none()

        qs = Region.objects.all()

        if self.q:
            qs = qs.filter(name__istartswith=self.q)

        return qs


class RanksAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return Rank.objects.none()

        qs = Rank.objects.all()

        if self.q:
            qs = qs.filter(name__istartswith=self.q)

        return qs


class TasksAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return Region.objects.none()

        qs = Task.objects.all()

        tasks_id = self.request.user.task.values_list('id', flat=True)

        if tasks_id:
            qs = qs.filter(id__in=tasks_id)

        if self.q:
            qs = qs.filter(name__icontains=self.q)

        return qs


class UsersAutocompleteView(autocomplete.Select2QuerySetView):

    def get_result_label(self, result):
        label = ''
        if result.fio.strip() != '':
            label += f'{result.fio} '
        if result.phone:
            label += f'{result.phone} '
        if result.email:
            label += f'{result.email} '
        if not label:
            label = f'Пользователь # {result.id}'
        return label

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return User.objects.none()

        qs = User.objects.all()

        if self.q:
            qs = qs.annotate(search_name=Concat('name', Value(' '), 'surname')).filter(
                Q(username__icontains=self.q) | Q(email__icontains=self.q) | Q(phone__icontains=self.q) | Q(
                    search_name__icontains=self.q))

        return qs


class UserSubsAutocompleteView(autocomplete.Select2QuerySetView):

    def get_result_label(self, result):
        label = ''
        if result.fio.strip() != '':
            label += f'{result.fio} '
        if result.phone:
            label += f'{result.phone} '
        if result.email:
            label += f'{result.email} '
        if not label:
            label = f'Супервайзер # {result.id}'
        return label

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return User.objects.none()

        qs = User.objects.filter(
            id__in=UserSub.objects.values('user_id').order_by('user_id').distinct('user_id')
        ).order_by('surname')
        if self.q:
            qs = qs.annotate(search_name=Concat('name', Value(' '), 'surname')).filter(
                Q(surname__icontains=self.q) | Q(email__icontains=self.q) | Q(
                    phone__icontains=self.q) | Q(name__icontains=self.q))
        return qs


class ExportView(FormView):
    template_name = 'survey/export.html'
    form_class = ExportForm
    success_url = reverse_lazy('survey:export')
    is_success = False

    def form_valid(self, form):

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="survey_export_%s.xls"' % \
                                          datetime.datetime.now()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Shop survey export data'))

        # Настройки
        start_row = 1
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        header = (
            ('Код клиента', 14),
            ('Код завода', 13),
            ('Регион', 14),
            ('Группа', 20),
            ('Клиент', 18),
            ('Адрес', 80),
            ('Широта', 12),
            ('Долгота', 12),
            ('Товар', 20),
            ('Товар Ид', 20),
        )

        items = StoreTask.objects.all(). \
            prefetch_related('store', 'store__client', 'store__region_o', 'store__category', 'task', 'only_user'). \
            order_by('store__code')

        # items = items.extra(
        #     select={
        #         'is_loyalty': 'SELECT count(agent_stores.id) FROM agent_stores WHERE agent_stores.loyalty_1c_code = '
        #                       'chl_stores.code'
        #     },
        # )

        if form.cleaned_data.get('tasks'):
            try:
                task = Task.objects.get(name=form.cleaned_data.get('tasks'))
                items = items.filter(task=task)
            except Task.DoesNotExist:
                pass

        if form.cleaned_data.get('categories'):
            try:
                category = Category.objects.get(name=form.cleaned_data.get('categories'))
                items = items.filter(store__category=category)
            except Category.DoesNotExist:
                pass

        if form.cleaned_data.get('regions'):
            try:
                region = Region.objects.get(name=form.cleaned_data.get('regions'))
                items = items.filter(store__region_o=region)
            except Region.DoesNotExist:
                pass

        tasks = items.order_by('task__name', 'task__id').values('task__name', 'task__id'). \
            annotate(dcount=Count('task__name'))

        header_2 = []
        for i in tasks:
            header_2.append(('Задание', 20))
            header_2.append(('Периодичность', 20))
            header_2.append(('Время', 20))
            header_2.append(('Email сюрвеера', 20))
            header_2.append(('Telegram канал', 20))

        header = list(header) + list(header_2)

        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        current_code = 'No code'
        row = start_row

        font = Font(name='Arial', size=8)

        for i in items:

            if current_code != i.store.code:

                current_code = i.store.code

                row += 1

                cell = ws.cell(row, 1)
                cell.value = i.store.code
                cell.font = font

                cell = ws.cell(row, 2)
                cell.value = i.store.factory_code if i.store and i.store.factory_code is not None else ''
                cell.font = font

                # if i.store.is_loyalty > 0:
                #     cell.comment = Comment('Этот магазин участвует в программе лояльности', 'Shop surver export')
                #     fill_column = 0
                #     for h in header:
                #         fill_column += 1
                #         fill_cell = ws.cell(row, fill_column)
                #         fill_cell.fill = PatternFill(fgColor=COLOR_INDEX[5], fill_type="solid")

                cell = ws.cell(row, 3)
                cell.value = str(i.store.region_o) if i.store.region_o else ''
                cell.font = font

                cell = ws.cell(row, 4)
                cell.value = str(i.store.category) if i.store.category else ''
                cell.font = font

                cell = ws.cell(row, 5)
                cell.value = str(i.store.client) if i.store.client else ''
                cell.font = font

                cell = ws.cell(row, 6)
                cell.value = str(i.store.address) if i.store.address else ''
                cell.font = font

                cell = ws.cell(row, 7)
                cell.value = str(i.store.latitude) if i.store.latitude else ''
                cell.font = font

                cell = ws.cell(row, 8)
                cell.value = str(i.store.longitude) if i.store.longitude else ''
                cell.font = font

                idx = 0

                for j in tasks:
                    column = idx * 5 + 11
                    idx += 1

                    cell = ws.cell(row, column)
                    cell.value = str(j['task__name']) if j['task__name'] else ''
                    cell.font = font

                ws.row_dimensions[row].height = 12

            column = 0
            idx = 0
            for j in tasks:
                if i.task.name == j['task__name']:
                    column = idx * 5 + 11
                idx += 1

            if column > 0:
                period = ''
                if i.is_once:
                    period = '*'
                if i.per_week:
                    period = '*%s' % i.per_week
                if i.days_of_week:
                    period = i.days_of_week.replace('.', ',')

                cell = ws.cell(row, column + 1)
                cell.value = str(period) if period else ''
                cell.font = font

                if i.hours_start and i.hours_end:
                    cell = ws.cell(row, column + 2)
                    cell.value = '%s-%s' % (i.hours_start, i.hours_end)
                    cell.font = font

                # Email
                cell = ws.cell(row, column + 3)
                cell.value = str(i.only_user.email) if i.only_user else ''
                cell.font = font

                # Telegram channel
                cell = ws.cell(row, column + 4)
                cell.value = str(i.telegram_channel_id) if i.telegram_channel_id else ''
                cell.font = font

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), 1 + start_row)
        ws.auto_filter.ref = filter_range

        wb.save(response)

        return response


class DownloadImagesCount(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(DownloadImagesCount, self).dispatch(request, *args, **kwargs)

    def get(self, request):

        try:
            datetime_start = self.request.GET.get('data_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        except (TypeError, ValueError):
            return HttpResponse('Неверная начальная дата')

        try:
            datetime_end = self.request.GET.get('data_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            return HttpResponse('Неверная конечная дата')

        images_count = TasksExecutionImage.objects.filter(
            task__date_end__gte=datetime_start, task__date_end__lte=datetime_end
        ).select_related('task')

        if self.request.GET.get('id_user'):
            images_count = images_count.filter(task__user__id=self.request.GET.get('id_user'))

        if self.request.GET.get('id_task'):
            images_count = images_count.filter(task__task__id=self.request.GET.get('id_task'))

        if self.request.GET.get('id_client'):
            images_count = images_count.filter(task__store__client__id=self.request.GET.get('id_client'))
        
        if self.request.GET.get('id_imagestep'):
            images_count = images_count.filter(
                constructor_step_name__icontains=TaskStep.objects.get(id=self.request.GET.get('id_imagestep')))

        # Partner
        tasks_id = self.request.user.task.values_list('id', flat=True)
        if tasks_id:
            images_count = images_count.filter(task__task__in=tasks_id)

        images_count = images_count.count()

        return HttpResponse(images_count)


class DownloadImages(FormView):

    form_class = UploadImageForm
    success_url = reverse_lazy('survey:download_images')
    is_success = False
    template_name = 'survey/images.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(DownloadImages, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        # Context
        context = super(DownloadImages, self).get_context_data(**kwargs)

        return context

    def post(self, request):

        try:
            datetime_start = self.request.POST.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        except (TypeError, ValueError):
            return HttpResponse('Ошибка. Неверная начальная дата')

        try:
            datetime_end = self.request.POST.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            return HttpResponse('Ошибка. Неверная конечная дата')

        try:
            images = TasksExecutionImage.objects.filter(
                task__date_end__gte=datetime_start, task__date_end__lte=datetime_end
            ).select_related('task', 'task__store', 'task__store__client')
        except:
            error_str = traceback.format_exc()
            return HttpResponse(f'<p><b>Ошибка взаимодействия с базой данных</b></p>{error_str}')

        file_name = self.request.POST.get('datetime_start') + '-' + self.request.POST.get('datetime_end')

        # Partner
        tasks_id = self.request.user.task.values_list('id', flat=True)
        if tasks_id:
            images = images.filter(task__task__in=tasks_id)

        if self.request.POST.get('user'):
            images = images.filter(task__user__id=self.request.POST.get('user'))
            file_name += '-user' + self.request.POST.get('user')

        if self.request.POST.get('task'):
            images = images.filter(task__task__id=self.request.POST.get('task'))
            file_name += '-task' + self.request.POST.get('task')

        if self.request.POST.get('client'):
            images = images.filter(task__store__client__id=self.request.POST.get('client'))
            file_name += '-client' + self.request.POST.get('client')

        if self.request.POST.get('imagestep'):
            images = images.filter(
                constructor_step_name__icontains=TaskStep.objects.get(id=self.request.POST.get('imagestep')))
            file_name += '-imagestep' + self.request.POST.get('imagestep')

        # Make zip
        path = settings.DOWNLOAD_PATH
        file_name += '-' + str(time.time()).replace('.', '')
        file_name_real = os.path.join(path, f'{file_name}')
        file_name_zip_file = f'{file_name}.zip'
        path = os.path.join(path, file_name)

        try:
            os.mkdir(path)
            os.chmod(path, 0o0777)

            for i in images:

                file_source = i.image.path
                if not os.path.isfile(file_source):
                    continue

                # Store
                store_dir_name = f'{i.task.store.client.name} {i.task.store.code}'.replace('/', '-')
                store_dir_name = store_dir_name[:100]
                store_dir = os.path.join(path, store_dir_name)
                try:
                    os.mkdir(store_dir)
                    os.chmod(store_dir, 0o0777)
                except OSError:
                    pass

                # Step
                if i.constructor_step_name:
                    step = i.constructor_step_name[:30]
                else:

                    steps = {
                        'undefined': 'Неизвестный шаг', 'enter': 'Вход магазина', 'before': 'Фото ДО',
                        'after': 'Фото ПОСЛЕ',
                        'check': 'Чек'
                    }
                    step = steps[i.type]

                step_dir = os.path.join(store_dir, step)
                try:
                    os.mkdir(step_dir)
                    os.chmod(step_dir, 0o0777)
                except OSError:
                    pass

                # Copy
                date_visit = i.task.date_end.strftime('%d.%m.%Y')
                source_file_name, source_file_extension = os.path.splitext(file_source)
                file_name = f'{store_dir_name}-{date_visit}-{step}-{i.id}{source_file_extension}'
                new_file = os.path.join(step_dir, file_name)
                shutil.copy2(file_source, new_file)

                # Watermark
                try:
                    img = Image.open(new_file)
                    image_width, image_height = img.size
                    font = ImageFont.truetype(
                        os.path.join(settings.STATIC_ROOT, 'admin/fonts/Roboto-Regular-webfont.woff'), 16)
                    draw = ImageDraw.Draw(img)

                    line_width_all, line_height_all = font.getsize(file_name)

                    text = i.task.store.client.name
                    line_width, line_height = font.getsize(text)
                    draw.text(
                        (image_width - line_width - 20, image_height - (line_height_all + 2) * 4 - 10),
                        text,
                        (159, 0, 0),
                        font=font
                    )
                    text = i.task.store.address
                    line_width, line_height = font.getsize(text)
                    draw.text(
                        (image_width - line_width - 20, image_height - (line_height_all + 2) * 3 - 10),
                        text,
                        (159, 0, 0),
                        font=font
                    )
                    text = date_visit
                    line_width, line_height = font.getsize(text)
                    draw.text(
                        (image_width - line_width - 20, image_height - (line_height_all + 2) * 2 - 10),
                        text,
                        (159, 0, 0),
                        font=font
                    )
                    text = step
                    line_width, line_height = font.getsize(text)
                    draw.text(
                        (image_width - line_width - 20, image_height - (line_height_all + 2) - 10),
                        text,
                        (159, 0, 0),
                        font=font
                    )
                    img.save(new_file)
                except:
                    pass

            # Zip
            shutil.make_archive(file_name_real, 'zip', path)
            file_name_real = f'{file_name_real}.zip'

            # Delete path
            shutil.rmtree(path)
        except:
            error_str = traceback.format_exc()
            try:
                shutil.rmtree(path)
            except:
                pass
            return HttpResponse(f'<p><b>Ошибка выгрузки файлов</b></p>{error_str}')
        
        try:
            with open(file_name_real, 'rb') as f:
                file_data = f.read()
        except FileNotFoundError:
            error_str = traceback.format_exc()
            return HttpResponse(f'<p><b>Файл не найден</b></p>{error_str}')
        except (OSError, IOError):
            error_str = traceback.format_exc()
            return HttpResponse(f'<p><b>Ошибка при открытии файла</b></p>{error_str}')

        response = HttpResponse(file_data, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename={file_name_zip_file}'

        return response


class ReportTeDeclinesView(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        def save_to_cell(ws, row, column, value):

            font = Font(name='Arial', size=8)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            cell = ws.cell(row, column)
            cell.value = value
            cell.font = font
            cell.border = border

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report_declines_%s.xls"' % \
                                          datetime.datetime.now()

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(_('Задачи с отказами'))

        # items
        items = TasksExecution.objects.filter(status=5).order_by('date_start')
        if datetime_start:
            items = items.filter(date_start__gte=datetime_start)
        if datetime_end:
            items = items.filter(date_start__lte=datetime_end)
        items = items.select_related('user', 'task', 'store', 'store__client')

        # Настройки
        start_row = 4
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        header = (
            ('Дата', 14),
            ('E-mail', 20),
            ('Название маршрута', 18),
            ('Задача', 20),
            ('Статус', 10),
            ('Код магазина', 18),
            ('Код завода', 17),
            ('Адрес', 50),
            ('Клиент ', 20),
            ('Комментарии аудитора ', 100),
        )

        # Title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
        cell = ws['A2']
        cell.value = 'Отказы по задачам {date_start} {date_end}'.format(
            date_start='с ' + datetime_start.strftime('%d.%m.%Y') if datetime_start else '',
            date_end='по ' + datetime_end.strftime('%d.%m.%Y') if datetime_end else '',
        )
        cell.font = Font(name='Arial', size=14, bold=True)
        rd = ws.row_dimensions[2]
        rd.height = 18

        # Rows
        row = start_row
        for i in items:
            row += 1

            rd = ws.row_dimensions[row]
            rd.height = 13

            save_to_cell(ws, row, 1, datetime.datetime.strftime(i.date_start, '%d.%m.%Y %H:%I:%S'))
            save_to_cell(ws, row, 2, i.user.email)
            save_to_cell(ws, row, 3, i.user.route)
            save_to_cell(ws, row, 4, i.task.name)
            save_to_cell(ws, row, 5, i.get_status_display())
            save_to_cell(ws, row, 6, i.store.code if i.store else '')
            save_to_cell(ws, row, 7, i.store.factory_code if i.store and i.store.factory_code else '')
            save_to_cell(ws, row, 8, i.store.address if i.store else '')
            save_to_cell(ws, row, 9, i.store.client.name if i.store else '')
            save_to_cell(ws, row, 10, i.comments_status)

        # Filters
        column = 0
        for i in header:
            column += 1
            cell = ws.cell(start_row, column)
            cell.value = i[0]
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.fill = PatternFill(start_color='f4ecc5', end_color='f4ecc5', fill_type='solid')
            ws.column_dimensions[get_column_letter(column)].width = i[1]

        filter_range = 'A%s:%s%s' % (start_row, get_column_letter(ws.max_column), items.count() + start_row)
        ws.auto_filter.ref = filter_range

        response.content = save_virtual_workbook(wb)
        return response


class ImportTasksView(FormView):
    template_name = 'survey/import-tasks.html'
    form_class = ImportTasksForm
    success_url = reverse_lazy('survey:import-tasks')
    is_success = False

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        if self.request.user.email not in settings.ADD_TASK_USERS:
            return HttpResponseForbidden()
        if self.request.GET.get('action') == 'cancel':
            cache.set('survey_import_tasks_cancel', 1)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):

        import hashlib
        import random

        try:
            ImportTask.objects.get(status=1)
            return
        except ImportTask.DoesNotExist:
            pass

        temp_file = '/tmp/%s.xls' % hashlib.sha1(str(random.random()).encode()).hexdigest()[:12]

        file_name = self.request.FILES['file'].name

        with open(temp_file, 'wb+') as destination:
            for chunk in self.request.FILES['file'].chunks():
                destination.write(chunk)
            destination.close()
            survey_import_tasks.delay(temp_file, file_name, self.request.user.id)

        messages.success(self.request, _('Import data task sending successfully'))

        return super().form_valid(form)

    def get_context_data(self, **kwargs):

        # Context
        context = super().get_context_data(**kwargs)

        # Search
        try:
            import_obj = ImportTask.objects.get(status=1)
        except ImportTask.DoesNotExist:
            import_obj = None

        # Messages
        storage = get_messages(self.request)
        for i in storage:
            self.is_success = True

        context['import_obj'] = import_obj
        context['is_success'] = self.is_success

        # Last import
        last_import = ImportTask.objects.first()
        context['last_import'] = last_import

        # Exit
        return context


class ReportAuditorsView(View):

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        try:
            datetime_start = request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
        except (TypeError, ValueError):
            datetime_start = None

        try:
            datetime_end = request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None

        if datetime_start is None or datetime_end is None:
            messages.error(self.request, 'Выберите дату дачала и дату завершения для отчета.')
            return redirect(reverse_lazy('survey:reports'))

        if (datetime_end - datetime_start).days > 31:
            messages.error(self.request, 'Максимальное количество дней отчета: 31')
            return redirect(reverse_lazy('survey:reports'))

        # items
        items = TasksExecution.objects.filter(
            Q(check_user__isnull=False) | Q(audit_user__isnull=False)
            # Q(check_type__in=['true', 'false']) | Q(is_auditor=True) Q(is_auditor=True) |
        ).order_by('date_start')
        if datetime_start:
            items = items.filter(date_start__gte=datetime_start)
        if datetime_end:
            items = items.filter(date_start__lte=datetime_end)
        items = items.select_related('user', 'task', 'store', 'store__client', 'check_user').values_list(
            'date_start', 'user__id', 'user__email', 'task__name', 'money', 'status', 'store__code',
            'store__address', 'check_type', 'audit_user__email', 'check_user__email',
        )

        # Excel
        output = io.BytesIO()
        workbook = Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        headers = ['Дата', 'Ид пользователя', 'Пользователь', 'Задача', 'Сумма, руб.', 'Статус', 'Код магазина',
                   'Адрес', 'Чек', 'Аудитор задачи', 'Аудитор чека/причин']
        widths = [15, 19, 28, 25, 15, 15, 17, 50, 18, 28, 28]

        header_format_dict = {'bg_color': '#eeeeee', 'bold': True, 'border': 1}
        h_fmt = workbook.add_format(header_format_dict)
        worksheet.write_row(0, 0, headers, h_fmt)
        worksheet.autofilter(0, 0, 10, 10)
        all_format_dict = {'font_size': 9}
        a_fmt = workbook.add_format(all_format_dict)
        format_dict = {'num_format': 'dd.mm.yyyy hh:mm:ss', 'font_size': 9}
        fmt = workbook.add_format(format_dict)
        worksheet.set_column(0, 0, cell_format=fmt)
        for i in range(len(widths)):
            worksheet.set_column(i, i, width=widths[i])

        # Make Excel data
        row_num = 0
        for i in items:
            row_num += 1
            col_num = 0
            for c in i:
                value = c
                if col_num == 5:
                    try:
                        v = {k: v for k, v in TasksExecution.statuses}[value]
                        value = str(v)
                    except:
                        pass
                if col_num == 8:
                    try:
                        v = {k: v for k, v in TasksExecution.check_types}[value]
                        value = str(v)
                    except:
                        pass
                # if col_num == 9:
                #     value = 'Да' if value else 'Нет'
                worksheet.write(row_num, col_num, value, a_fmt if col_num != 0 else fmt)
                col_num += 1

        # Done
        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="report_auditors_%s.xls"' % \
                                          datetime.datetime.now()
        output.close()

        return response


class ImageStepAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return TaskStep.objects.none()

        qs = TaskStep.objects.filter(step_type='photos').all()

        if self.q:
            qs = qs.filter(name__istartswith=self.q)

        task = self.forwarded.get('task', None)

        if task:
            qs = qs.filter(task=task)
        else:
            qs = TaskStep.objects.none()

        return qs


class UserStatusAutocompleteView(autocomplete.Select2QuerySetView):

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return UserStatus.objects.none()

        qs = UserStatus.objects.all()

        if self.q:
            qs = qs.filter(name__istartswith=self.q)

        return qs


class UsersDeviceAutocompleteView(autocomplete.Select2QuerySetView):

    def get_result_label(self, result):
        label = ''
        if result.fio.strip() != '':
            label += f'{result.fio} '
        if result.phone:
            label += f'{result.phone} '
        if result.email:
            label += f'{result.email} '
        if not label:
            label = f'Пользователь # {result.id}'
        return label

    def get_queryset(self):

        if not self.request.user.is_authenticated:
            return User.objects.none()

        qs = User.objects.filter(id__in=UserDevice.objects.values('user__id')).all()

        if self.q:
            qs = qs.annotate(search_name=Concat('name', Value(' '), 'surname')).filter(
                Q(username__icontains=self.q) | Q(email__icontains=self.q) | Q(phone__icontains=self.q) | Q(
                    search_name__icontains=self.q))

        return qs


class ReportDeviceSurveyorsView(TemplateView):
    """
    Отчет по версиям приложения (отчет №8)
    """
    template_name = 'survey/reports.devicesurveyors.html'

    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        return super(ReportDeviceSurveyorsView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # Context
        context = super(ReportDeviceSurveyorsView, self).get_context_data(**kwargs)

        try:
            page = self.request.GET.get('page', 1)
        except (TypeError, ValueError):
            page = 1

        try:
            datetime_start = self.request.GET.get('datetime_start')
            datetime_start = datetime.datetime.strptime(datetime_start, "%d.%m.%Y")
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        except (TypeError, ValueError):
            datetime_start = None
        try:
            datetime_end = self.request.GET.get('datetime_end')
            datetime_end = datetime.datetime.strptime(datetime_end, "%d.%m.%Y")
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        except (TypeError, ValueError):
            datetime_end = None
        if not datetime_start:
            datetime_start = datetime.datetime.today()
            datetime_start = datetime_start.replace(hour=0, minute=0, second=0, microsecond=0)
        if not datetime_end:
            datetime_end = datetime.datetime.today()
            datetime_end = datetime_end.replace(hour=0, minute=0, second=0, microsecond=0)
            datetime_end = datetime_end + datetime.timedelta(days=1) - datetime.timedelta(seconds=1)
        context['datetime_start'] = datetime_start
        context['datetime_start_d'] = datetime_start.strftime('%Y-%m-%d')
        context['datetime_end'] = datetime_end
        context['datetime_end_d'] = datetime_end.strftime('%Y-%m-%d')
        tasks_count = TasksExecution.objects.filter(
            user__id=OuterRef('id'),
            date_start__gte=datetime_start,
            date_end__lte=datetime_end
        ).values('user_id').annotate(
            count=Count('user_id')
        ).order_by().values('count')
        last_task_date = TasksExecution.objects.filter(
            user__id=OuterRef('id'),
            date_start__gte=datetime_start,
            date_end__lte=datetime_end
        ).values('user_id').annotate(
            max=Max('date_start')
        ).order_by().values('max')
        version = UserDevice.objects.filter(
            user__id=OuterRef('id')
        ).order_by('date_use').values('version')[:1]
        users_list = User.objects.filter().values(
            'id', 'name', 'surname', 'email', 'date_join'
        ).annotate(
            tasks_count=Subquery(tasks_count, output_field=IntegerField()),
            last_task_date=Subquery(last_task_date, output_field=DateField()),
            version=Subquery(version, output_field=CharField())
        ).filter(tasks_count__gt=0, version__isnull=False).order_by('version')
        content = []
        row = {}
        if_true = lambda x: x if x else ''
        if_true_zero = lambda x: x if x else 0
        for user in users_list:
            row = {
                'user_id': user['id'],
                'fio': User(name=user['name'], surname=user['surname']).fio + ', ' \
                       + if_true(user['email']),
                'date_join': user['date_join'],
                'te_max_date': user['last_task_date'],
                'te_count_done': str(if_true_zero(user['tasks_count'])),
                'version': user['version']
            }
            content.append(row)
        
        paginator = Paginator(content, 50)

        page_obj = paginator.get_page(page)
        context['page_obj'] = page_obj
        context['quantiy'] = len(users_list)

        data_r = []
        data_r.append({'data': page_obj})
        context['data'] = data_r

        # Exit
        return context

    def get(self, request, *args, **kwargs):
        response = super(ReportDeviceSurveyorsView, self).get(request, *args, **kwargs)
        if self.request.GET.get('export_report', None):

            context = self.get_context_data(*args, **kwargs)
            # Excel
            output = io.BytesIO()
            workbook = Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            all_format_dict = {'font_size': 9}
            a_fmt = workbook.add_format(all_format_dict)
            format_dict = {'num_format': 'dd.mm.yyyy hh:mm:ss', 'font_size': 9}
            d_fmt = workbook.add_format(format_dict)
            percent_fmt = workbook.add_format({'num_format': '0%'})
            headers = ['Пользователь', 'Дата регистрации',
                       'Дата последнего выполненного задания', 'Количество выполненных заданий',
                       'Версия приложения']
            widths = [45, 30, 30, 25, 25]
            # Make Excel data
            row_num = 0
            for i in context['data'][0]['data']:
                row_num += 1
                worksheet.write(row_num, 0, i['fio'], a_fmt)
                worksheet.write(row_num, 1, i['date_join'], d_fmt)
                worksheet.write(row_num, 2, i['te_max_date'], d_fmt)
                worksheet.write(row_num, 3, int(i['te_count_done']), a_fmt)
                worksheet.write(row_num, 4, i['version'], a_fmt)
            header_format_dict = {'bg_color': '#eeeeee', 'bold': True, 'border': 1}
            h_fmt = workbook.add_format(header_format_dict)
            worksheet.write_row(0, 0, headers, h_fmt)
            worksheet.autofilter(0, 0, 1, len(headers) - 1)
            for i in range(len(widths)):
                worksheet.set_column(i, i, width=widths[i])
            # Done
            workbook.close()
            output.seek(0)
            response = HttpResponse(output.read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename="report_device_%s.xls"' % \
                                              datetime.datetime.now()
            output.close()
        return response
